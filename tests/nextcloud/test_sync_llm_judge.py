"""
Tests for the LLM-as-judge feature in sync_llm.py: judge YAML parsing,
prompt assembly, label/logprob scoring, and the skip-already-scored xlsx
round-trip. Follows this test module's existing convention — no
Client/unittest.mock mocking, pure functions tested with plain inputs, real
tmp_path files for anything xlsx-shaped.
"""

import math

import pytest

from src.nextcloud.sync_llm import (
    _batch_row_for_transcript_stem,
    _build_judge_template,
    _locate_score_token,
    _missing_judge_cells,
    _normalize_anchors,
    _parse_judge_scale,
    _parse_judge_yaml,
    _read_judge_scores_xlsx,
    _render_judge_prompt,
    _scan_prompt_cells,
    _score_from_label,
    _score_from_logprobs,
    _write_judge_scores_xlsx,
)


class TestParseJudgeScale:
    """Test the numeric-range/discrete-set/label-list scale shapes."""

    def test_numeric_range_string(self):
        assert _parse_judge_scale("1-5") == {"type": "numeric", "values": [1, 2, 3, 4, 5]}

    def test_two_element_numeric_list_is_inclusive_range(self):
        assert _parse_judge_scale([1, 5]) == {"type": "numeric", "values": [1, 2, 3, 4, 5]}

    def test_discrete_numeric_list_preserved_as_is(self):
        assert _parse_judge_scale([1, 3, 5]) == {"type": "numeric", "values": [1, 3, 5]}

    def test_label_list(self):
        result = _parse_judge_scale(["poor", "fair", "good", "excellent"])
        assert result == {"type": "label", "values": ["poor", "fair", "good", "excellent"]}

    def test_invalid_scale_raises(self):
        with pytest.raises(ValueError):
            _parse_judge_scale(None)

    def test_empty_list_raises(self):
        with pytest.raises(ValueError):
            _parse_judge_scale([])


class TestNormalizeAnchors:
    """Test dict (explicit) vs. list (auto-distributed) anchor shapes."""

    def test_none_returns_empty(self):
        assert _normalize_anchors(None, {"type": "numeric", "values": [1, 2, 3, 4, 5]}) == {}

    def test_dict_used_as_is_sparse(self):
        scale = {"type": "numeric", "values": [1, 2, 3, 4, 5]}
        anchors = {1: "bad", 5: "great"}
        assert _normalize_anchors(anchors, scale) == {"1": "bad", "5": "great"}

    def test_list_distributes_across_numeric_scale_endpoints_inclusive(self):
        scale = {"type": "numeric", "values": list(range(1, 101))}
        result = _normalize_anchors(["worst", "mid", "best"], scale)
        # 100-element range has no single middle value; index round(99/2)=50
        # (Python's round-half-to-even) lands on value 51, not 50.
        assert result == {"1": "worst", "51": "mid", "100": "best"}

    def test_list_distributes_across_label_scale(self):
        scale = {"type": "label", "values": ["poor", "fair", "good", "excellent"]}
        result = _normalize_anchors(["bad", "great"], scale)
        assert result == {"poor": "bad", "excellent": "great"}

    def test_single_anchor_lands_in_the_middle(self):
        scale = {"type": "numeric", "values": [1, 2, 3, 4, 5]}
        assert _normalize_anchors(["average"], scale) == {"3": "average"}


class TestParseJudgeYaml:
    """Test required-field validation and the assembled shape."""

    VALID = """
kind: label
llm: qwen3-14b
jurisdiction: [summary]
scale: [poor, fair, good, excellent]
prompt: |
  Judge the summary quality.
"""

    def test_valid_yaml_produces_expected_shape(self):
        result = _parse_judge_yaml(self.VALID, "tone-quality")
        assert result["name"] == "tone-quality"
        assert result["llm"] == "qwen3-14b"
        assert result["kind"] == "label"
        assert result["jurisdiction"] == ["summary"]
        assert result["scale"] == {"type": "label", "values": ["poor", "fair", "good", "excellent"]}
        assert "{response}" in result["template"]

    def test_scalar_jurisdiction_normalized_to_list(self):
        text = self.VALID.replace("jurisdiction: [summary]", "jurisdiction: summary")
        result = _parse_judge_yaml(text, "j")
        assert result["jurisdiction"] == ["summary"]

    @pytest.mark.parametrize("field", ["prompt", "jurisdiction", "scale", "llm", "kind"])
    def test_missing_required_field_raises(self, field):
        data = {
            "prompt": "x",
            "jurisdiction": ["summary"],
            "scale": ["a", "b"],
            "llm": "qwen3-14b",
            "kind": "label",
        }
        del data[field]
        import yaml

        with pytest.raises(ValueError, match=field):
            _parse_judge_yaml(yaml.dump(data), "j")

    def test_invalid_kind_raises(self):
        text = self.VALID.replace("kind: label", "kind: nonsense")
        with pytest.raises(ValueError, match="kind"):
            _parse_judge_yaml(text, "j")


class TestBuildJudgeTemplate:
    """Test fixed section ordering and optional-section omission."""

    def _judge_data(self, **overrides):
        data = {
            "prompt": "Judge this.",
            "kind": "label",
            "_scale": {"type": "numeric", "values": [1, 2, 3, 4, 5]},
        }
        data.update(overrides)
        return data

    def test_section_order_with_all_optional_fields(self):
        data = self._judge_data(
            anchors={1: "bad", 5: "great"},
            few_shot=[{"response": "r1", "score": "3", "rationale": "ok"}],
            anti_pattern="Do not reward length.",
        )
        template = _build_judge_template(data)
        assert template.index("Judge this.") < template.index("Scale:")
        assert template.index("Scale:") < template.index("Anchors:")
        assert template.index("Anchors:") < template.index("Examples:")
        assert template.index("Examples:") < template.index("Do NOT reward:")
        assert template.index("Do NOT reward:") < template.index("Response to evaluate:")
        assert template.rstrip().endswith("only.") or "Respond with" in template

    def test_optional_sections_absent_when_field_omitted(self):
        template = _build_judge_template(self._judge_data())
        assert "Anchors:" not in template
        assert "Examples:" not in template
        assert "Do NOT reward:" not in template

    def test_label_vs_logprob_final_instruction_differs(self):
        label_template = _build_judge_template(self._judge_data(kind="label"))
        logprob_template = _build_judge_template(self._judge_data(kind="logprob"))
        assert "Respond with ONLY the label" in label_template
        assert "single token" in logprob_template


class TestRenderJudgePrompt:
    """Mirrors TestRenderPrompt for the judge-specific {response} placeholder."""

    def test_response_placeholder_substituted(self):
        assert _render_judge_prompt("Evaluate: {response}", "the output text") == "Evaluate: the output text"

    def test_transcript_and_context_substituted_when_referenced(self):
        result = _render_judge_prompt("{context}\n{transcript}\n{response}", "resp", "trans", "ctx")
        assert result == "ctx\ntrans\nresp"

    def test_unused_placeholders_are_a_no_op(self):
        assert _render_judge_prompt("Evaluate: {response}", "resp", "trans", "ctx") == "Evaluate: resp"

    def test_extra_field_placeholder_substituted(self):
        result = _render_judge_prompt(
            "Ground truth: {ground_truth}\n{response}", "resp", extra_fields={"ground_truth": "expected"}
        )
        assert result == "Ground truth: expected\nresp"


class TestScoreFromLabel:
    """Test label-reply parsing for both numeric and label scales."""

    def test_exact_case_insensitive_label_match(self):
        scale = {"type": "label", "values": ["poor", "fair", "good", "excellent"]}
        assert _score_from_label("GOOD", scale) == "good"

    def test_numeric_reply_parsed_directly(self):
        scale = {"type": "numeric", "values": [1, 2, 3, 4, 5]}
        assert _score_from_label("4", scale) == 4.0

    def test_numeric_fallback_regex_extracts_valid_value(self):
        scale = {"type": "numeric", "values": [1, 2, 3, 4, 5]}
        assert _score_from_label("I'd say a 4 out of 5.", scale) == 4.0

    def test_unparseable_numeric_returns_none(self):
        scale = {"type": "numeric", "values": [1, 2, 3, 4, 5]}
        assert _score_from_label("no idea", scale) is None

    def test_ambiguous_multi_label_substring_returns_none(self):
        scale = {"type": "label", "values": ["poor", "fair", "good", "excellent"]}
        assert _score_from_label("somewhere between fair and good", scale) is None

    def test_unambiguous_substring_fallback(self):
        scale = {"type": "label", "values": ["poor", "fair", "good", "excellent"]}
        assert _score_from_label("I think this is good overall.", scale) == "good"


class TestLocateScoreToken:
    """Test the post-</think> scan, including the false-positive guard for
    a pre-think token that coincidentally matches a scale value."""

    SCALE = {"type": "numeric", "values": [1, 2, 3, 4, 5]}

    def test_no_think_tag_scans_from_index_zero(self):
        logprobs = [{"token": "4", "top_logprobs": [{"token": "4", "logprob": -0.1}]}]
        assert _locate_score_token(logprobs, self.SCALE) == [{"token": "4", "logprob": -0.1}]

    def test_think_tag_split_across_tokens_still_advances(self):
        logprobs = [
            {"token": "<", "top_logprobs": []},
            {"token": "think", "top_logprobs": []},
            {"token": ">reasoning", "top_logprobs": []},
            {"token": "</think", "top_logprobs": []},
            {"token": ">", "top_logprobs": []},
            {"token": "4", "top_logprobs": [{"token": "4", "logprob": -0.2}]},
        ]
        assert _locate_score_token(logprobs, self.SCALE) == [{"token": "4", "logprob": -0.2}]

    def test_pre_think_coincidental_match_is_not_picked(self):
        # Reasoning body mentions "3 points" before </think> — must not be
        # picked over the real answer token that follows it.
        logprobs = [
            {"token": "<think>this covers", "top_logprobs": []},
            {"token": "3", "top_logprobs": [{"token": "3", "logprob": -5.0}]},
            {"token": " points</think>", "top_logprobs": []},
            {"token": "4", "top_logprobs": [{"token": "4", "logprob": -0.1}]},
        ]
        assert _locate_score_token(logprobs, self.SCALE) == [{"token": "4", "logprob": -0.1}]

    def test_no_match_anywhere_returns_none(self):
        logprobs = [{"token": "unsure", "top_logprobs": [{"token": "unsure", "logprob": -0.1}]}]
        assert _locate_score_token(logprobs, self.SCALE) is None


class TestScoreFromLogprobs:
    """Test the G-Eval probability-weighted average."""

    def test_numeric_weighted_average(self):
        scale = {"type": "numeric", "values": [1, 2, 3, 4, 5]}
        top_logprobs = [
            {"token": "4", "logprob": -0.05},
            {"token": "5", "logprob": -1.2},
            {"token": "3", "logprob": -3.8},
        ]
        score = _score_from_logprobs(top_logprobs, scale)
        w4, w5, w3 = math.exp(-0.05), math.exp(-1.2), math.exp(-3.8)
        expected = (4 * w4 + 5 * w5 + 3 * w3) / (w4 + w5 + w3)
        assert score == pytest.approx(expected)
        assert score == pytest.approx(4.219, abs=0.01)

    def test_label_scale_uses_ordinal_index(self):
        scale = {"type": "label", "values": ["poor", "fair", "good", "excellent"]}
        top_logprobs = [{"token": "good", "logprob": 0.0}]
        assert _score_from_logprobs(top_logprobs, scale) == pytest.approx(2.0)

    def test_zero_matching_candidates_returns_none(self):
        scale = {"type": "numeric", "values": [1, 2, 3, 4, 5]}
        top_logprobs = [{"token": "unsure", "logprob": -0.1}]
        assert _score_from_logprobs(top_logprobs, scale) is None


class TestScanPromptCells:
    """Regression coverage for the function extracted out of
    _build_prompt_tables (shared by the review-table builder and the judge
    pass)."""

    class _FakeClient:
        def __init__(self, listings):
            self._listings = listings

        def list(self, path, get_info=False):
            if path not in self._listings:
                from webdav3.exceptions import RemoteResourceNotFound

                raise RemoteResourceNotFound(path)
            return self._listings[path]

    def test_only_existing_output_cells_included(self):
        client = self._FakeClient(
            {
                "folder/transcriptions/": ["demo_a.txt", "demo_a.srt"],
                "folder/llm/": ["demo_a_glm-4-7-flash_summary.md"],
            }
        )
        batch_rows = {"demo_a": {"llm": "glm-4-7-flash", "prompt": "summary"}}
        cells = _scan_prompt_cells(client, "folder", batch_rows)
        assert cells == {"summary": {"demo_a": {"glm-4-7-flash": "demo_a_glm-4-7-flash_summary.md"}}}

    def test_missing_output_excluded(self):
        client = self._FakeClient(
            {
                "folder/transcriptions/": ["demo_a.txt"],
                "folder/llm/": [],
            }
        )
        batch_rows = {"demo_a": {"llm": "glm-4-7-flash", "prompt": "summary"}}
        assert _scan_prompt_cells(client, "folder", batch_rows) == {}

    def test_missing_transcript_dir_yields_no_cells(self):
        client = self._FakeClient({})
        batch_rows = {"demo_a": {"llm": "glm-4-7-flash", "prompt": "summary"}}
        assert _scan_prompt_cells(client, "folder", batch_rows) == {}


class TestReadWriteJudgeScoresXlsx:
    """Round-trip via a real workbook written to tmp_path.

    One judge's workbook now covers every jurisdiction prompt at once (one
    sheet per prompt + an "Overview" comparison sheet) instead of one file
    per (judge, prompt) — this is what makes comparing a judge's take
    across prompts possible without hunting across scattered files.
    """

    def test_round_trip_excludes_blank_cells(self, tmp_path):
        prompts_data = {"summary": {"demo_a": {"glm-4-7-flash": 4.2}, "demo_b": {}}}
        path = tmp_path / "tone_scores.xlsx"
        _write_judge_scores_xlsx(path, prompts_data)
        result = _read_judge_scores_xlsx(path)
        assert result == {"summary": {"demo_a": {"glm-4-7-flash": 4.2}}}

    def test_multiple_prompts_get_separate_sheets(self, tmp_path):
        prompts_data = {
            "summary": {"demo_a": {"m1": 4.0}},
            "rootcause": {"demo_a": {"m1": 2.0}},
        }
        path = tmp_path / "tone_scores.xlsx"
        _write_judge_scores_xlsx(path, prompts_data)
        result = _read_judge_scores_xlsx(path)
        assert result == {
            "summary": {"demo_a": {"m1": 4.0}},
            "rootcause": {"demo_a": {"m1": 2.0}},
        }

    def test_overview_sheet_not_treated_as_a_prompt_on_read(self, tmp_path):
        path = tmp_path / "tone_scores.xlsx"
        _write_judge_scores_xlsx(path, {"summary": {"demo_a": {"m1": 3.0}}})
        result = _read_judge_scores_xlsx(path)
        assert "Overview" not in result

    def test_overview_sheet_has_per_prompt_average(self, tmp_path):
        from openpyxl import load_workbook

        prompts_data = {"summary": {"demo_a": {"m1": 4.0}, "demo_b": {"m1": 2.0}}}
        path = tmp_path / "tone_scores.xlsx"
        _write_judge_scores_xlsx(path, prompts_data)
        wb = load_workbook(path, data_only=True)
        overview = list(wb["Overview"].iter_rows(values_only=True))
        assert overview[0] == ("prompt", "m1")
        assert overview[1] == ("summary", 3.0)

    def test_label_string_scores_dont_crash_overview_average(self, tmp_path):
        # kind: label on a non-numeric scale (e.g. poor/fair/good/excellent)
        # stores the label string itself as the score — found live crashing
        # the Overview sheet's averaging with
        # "TypeError: unsupported operand type(s) for +: 'int' and 'str'".
        prompts_data = {"summary": {"demo_a": {"m1": "good"}, "demo_b": {"m1": "excellent"}}}
        path = tmp_path / "tone_scores.xlsx"
        _write_judge_scores_xlsx(path, prompts_data)  # must not raise
        result = _read_judge_scores_xlsx(path)
        assert result == {"summary": {"demo_a": {"m1": "good"}, "demo_b": {"m1": "excellent"}}}

    def test_overview_averages_only_numeric_scores(self, tmp_path):
        from openpyxl import load_workbook

        # Mixed numeric (kind: logprob) and label-string (kind: label)
        # models in the same prompt — only the numeric one gets an average.
        prompts_data = {"summary": {"demo_a": {"numeric_judge": 4.0, "label_judge": "good"}}}
        path = tmp_path / "scores.xlsx"
        _write_judge_scores_xlsx(path, prompts_data)
        wb = load_workbook(path, data_only=True)
        rows = list(wb["Overview"].iter_rows(values_only=True))
        header, data_row = rows[0], dict(zip(rows[0][1:], rows[1][1:]))
        assert header[1:] == ("label_judge", "numeric_judge")
        assert data_row["numeric_judge"] == 4.0
        assert data_row["label_judge"] is None

    def test_charts_present_per_prompt_sheet_and_overview(self, tmp_path):
        from openpyxl import load_workbook

        prompts_data = {"summary": {"demo_a": {"m1": 3.0}}, "rootcause": {"demo_a": {"m1": 4.0}}}
        path = tmp_path / "tone_scores.xlsx"
        _write_judge_scores_xlsx(path, prompts_data)
        wb = load_workbook(path)
        assert len(wb["summary"]._charts) == 1
        assert len(wb["rootcause"]._charts) == 1
        assert len(wb["Overview"]._charts) == 1


class TestMissingJudgeCells:
    """The load-bearing test for the 'don't re-score' requirement."""

    def test_all_new_when_nothing_scored_yet(self):
        cells = {"demo_a": {"m1": "demo_a_m1_summary.md", "m2": "demo_a_m2_summary.md"}}
        assert set(_missing_judge_cells(cells, {})) == {
            ("demo_a", "m1", "demo_a_m1_summary.md"),
            ("demo_a", "m2", "demo_a_m2_summary.md"),
        }

    def test_partially_scored_stem_only_missing_model_returned(self):
        cells = {"demo_a": {"m1": "f1.md", "m2": "f2.md"}}
        existing = {"demo_a": {"m1": 4.0}}
        assert _missing_judge_cells(cells, existing) == [("demo_a", "m2", "f2.md")]

    def test_fully_scored_returns_empty(self):
        cells = {"demo_a": {"m1": "f1.md"}}
        existing = {"demo_a": {"m1": 4.0}}
        assert _missing_judge_cells(cells, existing) == []


class TestBatchRowForTranscriptStem:
    """Test mapping a transcript stem (possibly STT-suffixed) back to its
    original batch row."""

    def test_bare_stem_matches_directly(self):
        batch_rows = {"demo_a": {"llm": "glm-4-7-flash"}}
        assert _batch_row_for_transcript_stem("demo_a", batch_rows) == {"llm": "glm-4-7-flash"}

    def test_stt_suffixed_stem_matches_via_stt_field(self):
        batch_rows = {"demo_b": {"stt": ["turbo", "qwen3-asr-1.7b"], "context": "ctx"}}
        assert _batch_row_for_transcript_stem("demo_b_turbo", batch_rows)["context"] == "ctx"

    def test_no_match_returns_none(self):
        assert _batch_row_for_transcript_stem("unknown", {"demo_a": {}}) is None
