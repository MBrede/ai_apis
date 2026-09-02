"""
Nextcloud audio/video transcription sync.

Scans one or more configured Nextcloud folders (recursively, comma-separated
in NEXTCLOUD_FOLDER) for new audio/video files, transcribes them with speaker
diarization via the local Whisper API, and uploads .txt and .srt outputs to a
'transcriptions/' subfolder next to each source file.

Every audio/video file is transcribed unconditionally with the
WHISPER_MODEL-env default (unsuffixed output), UNLESS a `<stem>.yaml`/`.yml`
sidecar sits next to it with an `stt:` field (single model name or a list) —
in that case the file is transcribed once per listed model, with output
suffixed `<stem>_<model>.{txt,srt}` and skipped independently per model. See
README_llm.md for the full sidecar schema (shared with sync_llm.py).

A folder-level `batch.csv`/`batch.xlsx` (one row per file, same fields as
the per-file YAML sidecar: `filename`, `stt`, `llm`, `prompt`, `context`,
list-valued fields comma-separated within a cell) covers every file in that
folder that has no `<stem>.yaml` of its own — a customer-facing batch
alternative to hand-authoring one YAML per file. A file's own sidecar
always wins over a matching batch row.

All configuration is read from environment variables (see README_cron.md).
"""

import asyncio
import datetime
import json
import logging
import os
import re
import shutil
import tempfile
import warnings
from pathlib import Path

import aiohttp
import yaml
from dotenv import load_dotenv
from src.core.auth import build_auth_headers
from tqdm import tqdm
from webdav3.client import Client
from webdav3.exceptions import RemoteResourceNotFound

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
)
logger = logging.getLogger(__name__)

AUDIO_VIDEO_MIME_TYPES: frozenset[str] = frozenset(
    [
        "audio/wav",
        "audio/x-wav",
        "audio/mpeg",
        "audio/mp3",
        "audio/ogg",
        "audio/x-m4a",
        "audio/mp4",
        "audio/flac",
        "audio/aac",
        "audio/webm",
        "video/mp4",
        "video/x-msvideo",
        "video/quicktime",
        "video/x-matroska",
        "video/webm",
        "video/avi",
    ]
)

TRANSCRIPT_SUBFOLDER = "transcriptions"
SIDECAR_SUFFIXES = (".yaml", ".yml")
BATCH_FILENAMES = ("batch.xlsx", "batch.csv")  # checked in this order; first match wins
BATCH_ROW_LIST_FIELDS = ("stt", "llm", "prompt")
QWEN_MODEL_PREFIX = "qwen3-asr"
ARK_MODEL_PREFIX = "ark-asr"
HOJO_MODEL_PREFIX = "hojo-asr"
GRANITE_MODEL_PREFIX = "granite-speech"
NEMOTRON_MODEL_PREFIX = "nemotron"


def _normalize_list(value: object) -> list[str]:
    """Normalize a sidecar field that may be a scalar, a list, or absent.

    Args:
        value: Raw value from a parsed sidecar YAML dict (``data.get(key)``).

    Returns:
        ``[]`` if `value` is ``None``; the list of stringified items if
        `value` is already a list; a 1-element list of `str(value)` otherwise.
    """
    if value is None:
        return []
    if isinstance(value, list):
        return [str(v) for v in value]
    return [str(value)]


# ---------------------------------------------------------------------------
# Formatting helpers (mirrors bot.py output)
# ---------------------------------------------------------------------------


def _seconds_to_srt_timestamp(seconds: float) -> str:
    """Convert seconds to SRT timestamp HH:MM:SS,mmm."""
    total_ms = int(seconds * 1000)
    ms = total_ms % 1000
    total_s = total_ms // 1000
    h = total_s // 3600
    m = (total_s % 3600) // 60
    s = total_s % 60
    return f"{h:02d}:{m:02d}:{s:02d},{ms:03d}"


def _format_as_srt(segments: list[dict]) -> str:
    """Format diarization segments as SRT subtitles."""
    blocks = []
    for i, seg in enumerate(segments, start=1):
        start_ts = _seconds_to_srt_timestamp(seg["START"])
        end_ts = _seconds_to_srt_timestamp(seg["START"] + seg["DURATION"])
        blocks.append(
            f"{i}\n{start_ts} --> {end_ts}\n{seg['SPEAKER']}: {seg['TRANSCRIPTION'].strip()}"
        )
    return "\n\n".join(blocks)


def _format_as_text(segments: list[dict]) -> str:
    """Format diarization segments as a readable plain-text transcript."""
    lines = []
    for seg in segments:
        start_str = str(datetime.timedelta(seconds=int(seg["START"])))
        end_str = str(datetime.timedelta(seconds=int(seg["START"] + seg["DURATION"])))
        lines.append(
            f"[{start_str} - {end_str}] {seg['SPEAKER']}: {seg['TRANSCRIPTION'].strip()}"
        )
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# WebDAV helpers
# ---------------------------------------------------------------------------


def _make_webdav_client() -> tuple[Client, list[str]]:
    """Create a WebDAV client from environment variables.

    NEXTCLOUD_FOLDER may list several folders, comma-separated (e.g.
    ``"transcription,/Shared/1 Projects/MBS Benkana/transcription"``) — each
    is scanned independently by callers. A single folder (no comma) produces
    a 1-element list, identical to the pre-multi-folder behavior.

    Returns:
        Tuple of (client, list of remote root paths, one per folder).

    Raises:
        KeyError: If a required environment variable is missing.
    """
    url = os.environ["NEXTCLOUD_URL"].rstrip("/")
    user = os.environ["NEXTCLOUD_USER"]
    password = os.environ["NEXTCLOUD_PASSWORD"]
    folders = [f.strip().strip("/") for f in os.environ["NEXTCLOUD_FOLDER"].split(",") if f.strip()]
    # NEXTCLOUD_DAV_USER: the internal Nextcloud username used in the WebDAV path.
    # Defaults to NEXTCLOUD_USER. Set this when the login credential (email) differs
    # from the internal username, e.g. NEXTCLOUD_USER=user@example.com but
    # WebDAV path uses /remote.php/dav/files/username/.
    dav_user = os.environ.get("NEXTCLOUD_DAV_USER") or user
    client = Client(
        {
            "webdav_hostname": url,
            "webdav_login": user,
            "webdav_password": password,
        }
    )
    remote_roots = [f"/remote.php/dav/files/{dav_user}/{folder}" for folder in folders]
    return client, remote_roots


def _download_text(client: Client, remote_path: str, tmp_dir: Path) -> str:
    """Download a small remote text file and return its contents.

    Shared by sync.py (sidecar `stt:` parsing) and sync_llm.py (sidecar
    parsing, prompt templates, transcripts).

    Args:
        client: WebDAV client.
        remote_path: Remote file path to download.
        tmp_dir: Local scratch directory for the temporary download.

    Returns:
        Decoded UTF-8 file contents.
    """
    local_path = tmp_dir / Path(remote_path).name
    client.download_sync(remote_path=remote_path, local_path=str(local_path))
    try:
        return local_path.read_text(encoding="utf-8")
    finally:
        local_path.unlink(missing_ok=True)


def _stt_models_from_dict(data: dict) -> list[str]:
    """Return the STT models requested by a parsed sidecar/batch-row dict's `stt` field.

    Shared core of `_parse_sidecar_stt` (YAML sidecar) and the batch-file
    path (`_load_batch_rows` already produces this dict shape directly, no
    YAML round-trip needed).

    Returns:
        List of requested model names. Empty means "no opt-in" — callers
        should fall back to the WHISPER_MODEL-env default with unsuffixed
        output (today's behavior).
    """
    return _normalize_list(data.get("stt"))


def _parse_sidecar_stt(text: str) -> list[str]:
    """Return the STT models requested by a sidecar's `stt:` field.

    Args:
        text: Raw sidecar YAML contents.

    Returns:
        List of requested model names. Empty means "no sidecar opt-in" —
        callers should fall back to the WHISPER_MODEL-env default with
        unsuffixed output (today's behavior).
    """
    return _stt_models_from_dict(yaml.safe_load(text) or {})


BATCH_ROW_RESERVED_FIELDS = frozenset({"filename", "file", "context", *BATCH_ROW_LIST_FIELDS})


def _row_to_sidecar_dict(row: dict[str, str]) -> dict:
    """Convert one batch.csv/batch.xlsx row into the same dict shape
    `yaml.safe_load()` produces for an equivalent per-file YAML sidecar.

    Args:
        row: Column name -> cell value (both already lowercased/stripped by
            the caller). `stt`/`llm`/`prompt` are comma-split into lists;
            `context` is passed through as free text. Any OTHER column
            (e.g. a user-added `ground_truth` column) is passed through
            as-is too — sync_llm.py's `_sidecar_from_dict` exposes these as
            `extra_fields`, usable as `{field_name}` placeholders in both
            prompt and judge templates. Same passthrough happens naturally
            for a YAML sidecar's own extra top-level keys, since
            `yaml.safe_load` already keeps them — this function only needs
            to not drop them for the batch-row path.

    Returns:
        Dict with only the keys that had a non-empty cell — same convention
        `_stt_models_from_dict`/sync_llm.py's `_sidecar_from_dict` already
        expect from `data.get(key)` returning `None` for an absent field.
    """
    result: dict = {}
    for field in BATCH_ROW_LIST_FIELDS:
        value = (row.get(field) or "").strip()
        if value:
            result[field] = [v.strip() for v in value.split(",") if v.strip()]
    context = (row.get("context") or "").strip()
    if context:
        result["context"] = context
    for key, value in row.items():
        if key in BATCH_ROW_RESERVED_FIELDS:
            continue
        # Always include the column, even blank — `_read_batch_xlsx`/
        # `_read_batch_csv` already give every row a key per header, so
        # this just preserves that. Skipping blanks here would mean a
        # {field_name} placeholder referencing this column stays literally
        # unreplaced for any row/file that happens to leave the cell empty
        # (not substituted with "" the way {context} is), which is a real
        # bug found live — a judge's {ground_truth} placeholder showed up
        # unrendered for the two demo files that had no ground_truth cell.
        result[key] = (value or "").strip()
    return result


def _load_batch_rows(client: Client, files: list[dict], remote_dir: str, tmp_dir: Path) -> dict[str, dict]:
    """Find and parse a folder-level batch.csv/batch.xlsx, if present.

    One row per file (matched by a `filename` column against each audio
    file's stem, extension optional), same fields as the per-file YAML
    sidecar. Lets a customer hand over one spreadsheet instead of
    authoring a `<stem>.yaml` per file — see this module's docstring.

    Args:
        client: WebDAV client.
        files: File-info dicts for this directory (from `client.list(..., get_info=True)`).
        remote_dir: Directory path these files live in (for logging only).
        tmp_dir: Local scratch directory for the temporary download.

    Returns:
        `{stem: row_dict}`, `row_dict` in the same shape `_row_to_sidecar_dict`
        returns. Empty dict if no batch file is present in this folder.
    """
    batch_file = next(
        (f for name in BATCH_FILENAMES for f in files if Path(f["path"]).name.lower() == name),
        None,
    )
    if batch_file is None:
        return {}

    local_path = tmp_dir / Path(batch_file["path"]).name
    client.download_sync(remote_path=batch_file["path"], local_path=str(local_path))
    try:
        suffix = local_path.suffix.lower()
        if suffix == ".xlsx":
            rows = _read_batch_xlsx(local_path)
        else:
            rows = _read_batch_csv(local_path)
    except Exception as exc:
        logger.error("Failed to read batch file %s in %s: %s", batch_file["path"], remote_dir, exc)
        return {}
    finally:
        local_path.unlink(missing_ok=True)

    batch_rows: dict[str, dict] = {}
    for row in rows:
        filename = (row.get("filename") or row.get("file") or "").strip()
        if not filename:
            continue
        batch_rows[Path(filename).stem] = _row_to_sidecar_dict(row)
    return batch_rows


def _read_batch_csv(local_path: Path) -> list[dict[str, str]]:
    """Read a batch.csv into a list of {lowercased column name: cell value} dicts."""
    import csv

    with open(local_path, newline="", encoding="utf-8-sig") as fh:
        reader = csv.DictReader(fh)
        return [{(k or "").strip().lower(): (v or "") for k, v in row.items()} for row in reader]


def _read_batch_xlsx(local_path: Path) -> list[dict[str, str]]:
    """Read a batch.xlsx into a list of {lowercased column name: cell value} dicts."""
    from openpyxl import load_workbook

    wb = load_workbook(local_path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = [str(h or "").strip().lower() for h in next(rows_iter)]
    except StopIteration:
        return []
    return [
        {headers[i]: ("" if v is None else str(v)) for i, v in enumerate(row) if i < len(headers)}
        for row in rows_iter
        if any(v is not None for v in row)
    ]


def _backend_for_stt_model(model: str) -> str:
    """Map a sidecar-requested STT model name to a whisper API `backend` value.

    Args:
        model: Model name from a sidecar's `stt:` field (e.g. "turbo",
            "qwen3-asr-1.7b").

    Returns:
        "qwen3-asr"/"ark-asr"/"hojo-asr"/"granite-speech"/"nemotron-asr" if
        the model name starts with the matching prefix, else "whisperx"
        (today's implicit default backend).
    """
    name = model.lower()
    if name.startswith(QWEN_MODEL_PREFIX):
        return "qwen3-asr"
    if name.startswith(ARK_MODEL_PREFIX):
        return "ark-asr"
    if name.startswith(HOJO_MODEL_PREFIX):
        return "hojo-asr"
    if name.startswith(GRANITE_MODEL_PREFIX):
        return "granite-speech"
    if name.startswith(NEMOTRON_MODEL_PREFIX):
        return "nemotron-asr"
    return "whisperx"


def _ensure_transcript_folder(client: Client, remote_dir: str) -> str:
    """Return the transcriptions subfolder path, creating it if needed.

    Args:
        client: WebDAV client.
        remote_dir: Remote directory that contains the source audio file.

    Returns:
        Remote path of the transcriptions subfolder (trailing slash included).
    """
    transcript_dir = remote_dir.rstrip("/") + f"/{TRANSCRIPT_SUBFOLDER}/"
    try:
        client.list(transcript_dir)
    except RemoteResourceNotFound:
        client.mkdir(transcript_dir)
    return transcript_dir


def _collect_new_files(client: Client, remote_roots: list[str], tmp_dir: Path) -> list[dict]:
    """Recursively find audio/video work items that still need transcription.

    Each returned item is one (file, STT model) pair. A file with no sidecar
    and no matching batch.csv/batch.xlsx row, or a sidecar/row with no
    `stt:`/`stt` field, yields exactly one item with `stt_model=None` —
    today's behavior: single WHISPER_MODEL-env-default transcription,
    unsuffixed `<stem>.{txt,srt}` output, skipped if `<stem>` already has a
    transcript. A file with `stt: [m1, m2, ...]` (own sidecar, or a batch
    row if it has no sidecar of its own — see `_load_batch_rows`) yields one
    item per listed model, each independently skippable via its own
    `<stem>_<model>.{txt,srt}` output.

    Args:
        client: WebDAV client.
        remote_roots: Root remote folders to scan (see _make_webdav_client).
        tmp_dir: Local scratch directory for downloading sidecars.

    Returns:
        List of dicts: {"file": <WebDAV file-info dict>, "stt_model": str | None}.
    """
    work_items: list[dict] = []

    def _walk(path: str) -> None:
        try:
            entries = client.list(path, get_info=True)
        except RemoteResourceNotFound:
            logger.warning("Remote path not found, skipping: %s", path)
            return

        subdirs = [e for e in entries if e["isdir"] and e["path"].rstrip("/") != path.rstrip("/")]
        files = [e for e in entries if not e["isdir"]]
        audio_video = [f for f in files if f.get("content_type", "") in AUDIO_VIDEO_MIME_TYPES]

        if audio_video:
            audio_stems = {Path(f["path"]).stem for f in audio_video}
            sidecars_by_stem = {
                Path(f["path"]).stem: f["path"]
                for f in files
                if Path(f["path"]).suffix.lower() in SIDECAR_SUFFIXES and Path(f["path"]).stem in audio_stems
            }
            batch_rows = _load_batch_rows(client, files, path, tmp_dir)

            transcript_dir = path.rstrip("/") + f"/{TRANSCRIPT_SUBFOLDER}/"
            try:
                existing_stems = {Path(f).stem for f in client.list(transcript_dir)}
            except RemoteResourceNotFound:
                existing_stems = set()

            for f in audio_video:
                stem = Path(f["path"]).stem
                sidecar_path = sidecars_by_stem.get(stem)
                stt_models: list[str] = []
                if sidecar_path:
                    try:
                        text = _download_text(client, sidecar_path, tmp_dir)
                        stt_models = _parse_sidecar_stt(text)
                    except Exception as exc:
                        logger.error("Failed to read sidecar %s: %s", sidecar_path, exc)
                elif stem in batch_rows:
                    # No per-file sidecar — fall back to this folder's batch.csv/batch.xlsx row.
                    stt_models = _stt_models_from_dict(batch_rows[stem])

                if not stt_models:
                    # No sidecar, or sidecar has no `stt:` — today's default behavior.
                    if stem not in existing_stems:
                        work_items.append({"file": f, "stt_model": None})
                else:
                    for model in stt_models:
                        if f"{stem}_{model}" not in existing_stems:
                            work_items.append({"file": f, "stt_model": model})

        for d in subdirs:
            _walk(d["path"])

    for root in remote_roots:
        _walk(root)
    return work_items


# ---------------------------------------------------------------------------
# Transcription
# ---------------------------------------------------------------------------


def _fillers_from_filename(stem: str) -> bool:
    """Return True if the filename stem requests filler-word retention.

    Recognised patterns (case-insensitive): ``_filler`` or ``_fillers`` anywhere in the stem.

    Examples::

        interview_fillers_2.mp3  → True
        meeting_filler.wav       → True
        lecture_2.mp3            → False
    """
    return bool(re.search(r"(?:^|_)fillers?(?:_|$)", stem, re.IGNORECASE))


def _speakers_from_filename(stem: str) -> int | None:
    """Try to extract a speaker count encoded in a filename stem.

    Recognised patterns (case-insensitive, anywhere in the stem):
        interview_2         → trailing _<digits>
        session_2spk        → _<digits>spk[s]
        recording_2speaker  → _<digits>speaker[s]

    Args:
        stem: Filename without extension, e.g. "interview_2" or "session_2spk".

    Returns:
        Parsed speaker count, or None if no pattern matches.
    """
    patterns = [
        r"_(\d+)speakers?$",
        r"_(\d+)spks?$",
        r"_(\d+)$",
    ]
    for pat in patterns:
        m = re.search(pat, stem, re.IGNORECASE)
        if m:
            return int(m.group(1))
    return None


def _build_diarize_params(
    stem: str | None = None, model_override: str | None = None
) -> dict[str, str | int | bool]:
    """Build query parameters for the /transcribe_and_diarize/ endpoint.

    Speaker count priority:
      1. Encoded in the filename stem (e.g. ``interview_2.mp3`` → 2 speakers)
      2. ``NUM_SPEAKERS`` environment variable
      3. ``MIN_SPEAKERS`` + ``MAX_SPEAKERS`` environment variables

    Filler-word retention is enabled when the filename stem contains ``_filler``
    or ``_fillers`` (e.g. ``interview_fillers_2.mp3``).

    Args:
        stem: Filename stem of the file being transcribed (without extension).
        model_override: STT model requested via a sidecar's `stt:` field.
            When set, also selects the matching `backend` (see
            `_backend_for_stt_model`) instead of the endpoint's default.
            `None` preserves today's exact behavior (WHISPER_MODEL env
            default, implicit `whisperx` backend).

    Returns:
        Query-parameter dict for the diarization endpoint.

    Raises:
        ValueError: If no speaker count can be determined.
    """
    params: dict[str, str | int | bool] = {
        "model_to_use": model_override or os.environ.get("WHISPER_MODEL", "turbo"),
    }
    if model_override is not None:
        params["backend"] = _backend_for_stt_model(model_override)

    # Filler-word retention from filename
    if stem is not None and _fillers_from_filename(stem):
        logger.info("Filler-word retention enabled from filename '%s'.", stem)
        params["include_fillers"] = True

    # 1. Filename-encoded speaker count
    if stem is not None:
        n = _speakers_from_filename(stem)
        if n is not None:
            logger.info("Using %d speaker(s) from filename '%s'.", n, stem)
            params["num_speakers"] = n
            return params

    # 2. Environment variables
    num = os.environ.get("NUM_SPEAKERS")
    min_s = os.environ.get("MIN_SPEAKERS")
    max_s = os.environ.get("MAX_SPEAKERS")
    if num:
        params["num_speakers"] = int(num)
    elif min_s and max_s:
        params["min_speakers"] = int(min_s)
        params["max_speakers"] = int(max_s)
    else:
        raise ValueError(
            "Cannot determine speaker count for this file. "
            "Encode it in the filename (e.g. interview_2.mp3) or set "
            "NUM_SPEAKERS / MIN_SPEAKERS + MAX_SPEAKERS in your environment."
        )
    return params


async def _transcribe_file(
    local_path: str,
    mime_type: str,
    params: dict[str, str | int],
    session: aiohttp.ClientSession,
) -> tuple[list[dict], float | None] | None:
    """POST a file to the Whisper diarization endpoint.

    Args:
        local_path: Path to the downloaded audio/video file.
        mime_type: MIME type reported by Nextcloud.
        params: Query parameters (speaker count, model, api_key).
        session: Shared aiohttp session.

    Returns:
        `(segments, audio_duration)` on success — `audio_duration` in
        seconds, `None` if the whisper API response predates this field
        (older deployed image) rather than a hard failure. `None` (not a
        tuple) on request failure.
    """
    whisper_url = os.environ.get("WHISPER_URL", "http://whisper:8080")
    endpoint = f"{whisper_url.rstrip('/')}/transcribe_and_diarize/"
    filename = Path(local_path).name

    try:
        with open(local_path, "rb") as fh:
            form = aiohttp.FormData()
            form.add_field("file", fh, filename=filename, content_type=mime_type)
            auth_headers = build_auth_headers(os.environ.get("WHISPER_API_KEY"))
            async with session.post(endpoint, params=params, data=form, headers=auth_headers) as resp:
                if not (200 <= resp.status < 400):
                    body = await resp.text()
                    warnings.warn(
                        f"Transcription failed for {filename} (HTTP {resp.status}): {body}"
                    )
                    return None
                body = json.loads(await resp.read())
                return body["answer"], body.get("audio_duration")
    except Exception as exc:
        warnings.warn(f"Error transcribing {filename}: {exc}")
        return None


# A transcript is only flagged if it falls short by BOTH an absolute margin
# and a proportional one — small trailing silence (someone leaving the mic
# on, applause) shouldn't trip this; a backend cutting off well before the
# real end of the file should.
TRUNCATION_MIN_GAP_SECONDS = 5.0
TRUNCATION_MIN_GAP_FRACTION = 0.1

# Max whisper transcription calls in flight at once, PER target container
# (see "container_group" below — the primary container and each of its
# qwen3-asr/hojo-asr sidecars get their own independent budget, since they're
# separate GPU-scaled deployments and shouldn't compete for the same slots).
# Lower than sync_llm.py's LLM_CONCURRENCY (default 10) on purpose — this
# hits actual GPU inference directly (no request-queuing serving layer like
# vLLM in front of it), so too much concurrency risks CUDA OOM rather than
# just queuing. Env-tunable per cluster GPU headroom.
WHISPER_CONCURRENCY = int(os.environ.get("WHISPER_CONCURRENCY", "3"))


def _detect_possible_truncation(segments: list[dict], audio_duration: float | None) -> str | None:
    """Flag a transcript whose last segment ends well before the real audio ends.

    A cheap, blunt signal for backend-side truncation (e.g. a generation-length
    cap cutting a chunk off early, or a diarized speaker turn fed to a model
    far outside its validated input length) — not a precise content check,
    and it can't catch truncation mid-file (only at the very end of the
    transcribed output). Found live: qwen3-asr's transformers-backend
    `max_new_tokens` defaulted to 512, badly truncating anything beyond a
    short clip (see `stt_backends/qwen3_asr.py`).

    Args:
        segments: This file's merged speaker segments (may be empty) — the
            same list `_format_as_text`/`_format_as_srt` render, with
            `"START"`/`"DURATION"` keys per `merge_speaker_segments`.
        audio_duration: Real audio duration in seconds, as reported by the
            whisper API (see `_audio_duration_seconds` in whisper_api.py),
            or `None` if the response predates that field (older deployed
            image) — no check is possible then, not treated as a failure.

    Returns:
        A human-readable "Obacht" message if truncation looks likely, else `None`.
    """
    if not audio_duration:
        return None
    last_end = max((s["START"] + s["DURATION"] for s in segments), default=0.0)
    gap = audio_duration - last_end
    if gap > TRUNCATION_MIN_GAP_SECONDS and gap > audio_duration * TRUNCATION_MIN_GAP_FRACTION:
        return (
            f"Obacht: dieses Transkript ist evtl. abgeschnitten — die letzte Zeile endet bei "
            f"{last_end:.1f}s, aber die Audiodatei ist {audio_duration:.1f}s lang "
            f"({gap:.1f}s / {gap / audio_duration:.0%} fehlen am Ende)."
        )
    return None


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------


async def main() -> None:
    """Main sync loop: scan → download → transcribe → upload."""
    load_dotenv()

    client, remote_roots = _make_webdav_client()

    tmp_dir = Path(tempfile.mkdtemp(prefix="nc_transcribe_"))
    try:
        logger.info("Scanning %s for new audio/video files...", remote_roots)
        work_items = _collect_new_files(client, remote_roots, tmp_dir)

        if not work_items:
            logger.info("No new files to transcribe.")
            return

        logger.info("Found %d work item(s) to transcribe.", len(work_items))

        timeout = aiohttp.ClientTimeout(total=int(os.environ.get("WHISPER_TIMEOUT", "3600")))
        async with aiohttp.ClientSession(timeout=timeout) as session:
            # Downloads (webdav I/O, cheap) stay sequential; only the actual
            # whisper call is parallelized below, bounded by WHISPER_CONCURRENCY.
            # local_audio is index-prefixed rather than output_stem-derived —
            # two different source folders can share a filename (e.g. two
            # projects both containing "Neue Aufnahme 27.m4a"), which was only
            # safe under the old strictly-one-at-a-time loop; concurrent
            # transcription needs a guaranteed-unique local path per item.
            prepared = []
            for idx, item in enumerate(work_items):
                file_info: dict = item["file"]
                stt_model: str | None = item["stt_model"]
                remote_path: str = file_info["path"]
                mime_type: str = file_info.get("content_type", "audio/wav")
                stem = Path(remote_path).stem
                output_stem = f"{stem}_{stt_model}" if stt_model else stem
                suffix = Path(remote_path).suffix or ".wav"
                local_audio = tmp_dir / f"{idx}_{output_stem}{suffix}"

                try:
                    client.download_sync(remote_path=remote_path, local_path=str(local_audio))
                except Exception as exc:
                    logger.error("Download failed for %s: %s", remote_path, exc)
                    continue

                try:
                    diarize_params = _build_diarize_params(stem, model_override=stt_model)
                except ValueError as exc:
                    logger.error("Skipping %s: %s", remote_path, exc)
                    local_audio.unlink(missing_ok=True)
                    continue

                prepared.append(
                    {
                        "remote_path": remote_path,
                        "output_stem": output_stem,
                        "local_audio": local_audio,
                        "mime_type": mime_type,
                        "diarize_params": diarize_params,
                        # "qwen3-asr" and "hojo-asr" run in their own separate
                        # containers/GPU-scaled deployments (see
                        # whisper_api.py's module docstring) — everything else
                        # (whisperx/ark-asr/granite-speech/nemotron-asr) is
                        # handled by the primary container. Each group gets
                        # its own concurrency budget below so items destined
                        # for one container don't eat into another's queue.
                        "container_group": _backend_for_stt_model(stt_model) if stt_model else "whisperx",
                    }
                )

            proxied_groups = {"qwen3-asr", "hojo-asr"}
            semaphores: dict[str, asyncio.Semaphore] = {}

            def _semaphore_for(group: str) -> asyncio.Semaphore:
                key = group if group in proxied_groups else "primary"
                if key not in semaphores:
                    semaphores[key] = asyncio.Semaphore(WHISPER_CONCURRENCY)
                return semaphores[key]

            async def _run_transcribe(p: dict) -> tuple[dict, tuple[list[dict], float | None] | None]:
                async with _semaphore_for(p["container_group"]):
                    result = await _transcribe_file(str(p["local_audio"]), p["mime_type"], p["diarize_params"], session)
                p["local_audio"].unlink(missing_ok=True)
                return p, result

            tasks = [asyncio.ensure_future(_run_transcribe(p)) for p in prepared]
            for coro in tqdm(asyncio.as_completed(tasks), total=len(tasks), desc="Processing"):
                p, result = await coro
                if result is None:
                    continue
                segments, audio_duration = result
                remote_path = p["remote_path"]
                output_stem = p["output_stem"]

                # Write and upload outputs
                remote_dir = str(Path(remote_path).parent)
                transcript_dir = _ensure_transcript_folder(client, remote_dir)

                for content, ext in (
                    (_format_as_text(segments), ".txt"),
                    (_format_as_srt(segments), ".srt"),
                ):
                    local_out = tmp_dir / f"{output_stem}{ext}"
                    local_out.write_text(content, encoding="utf-8")
                    remote_out = transcript_dir + f"{output_stem}{ext}"
                    try:
                        client.upload_sync(local_path=str(local_out), remote_path=remote_out)
                        logger.info("Uploaded %s", remote_out)
                    except Exception as exc:
                        logger.error("Upload failed for %s: %s", remote_out, exc)
                    local_out.unlink(missing_ok=True)

                obacht = _detect_possible_truncation(segments, audio_duration)
                obacht_remote = transcript_dir + f"{output_stem}_OBACHT.txt"
                if obacht:
                    logger.warning("%s: %s", output_stem, obacht)
                    local_obacht = tmp_dir / f"{output_stem}_OBACHT.txt"
                    local_obacht.write_text(obacht, encoding="utf-8")
                    try:
                        client.upload_sync(local_path=str(local_obacht), remote_path=obacht_remote)
                    except Exception as exc:
                        logger.error("Upload failed for %s: %s", obacht_remote, exc)
                    local_obacht.unlink(missing_ok=True)
                else:
                    # Clear a stale flag from a previous run (e.g. re-run
                    # with a fixed backend/config) rather than leaving a
                    # now-inaccurate warning sitting next to a good transcript.
                    try:
                        client.clean(obacht_remote)
                    except RemoteResourceNotFound:
                        pass
                    except Exception as exc:
                        logger.error("Failed to clear stale flag %s: %s", obacht_remote, exc)

    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)

    logger.info("Done.")


if __name__ == "__main__":
    asyncio.run(main())
