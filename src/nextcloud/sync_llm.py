"""
Nextcloud transcript LLM post-processing.

Scans one or more configured Nextcloud folders (recursively, comma-separated
in NEXTCLOUD_FOLDER) for audio/video files that have an optional
`<stem>.yaml` sidecar next to them, OR a matching row in that folder's
`batch.csv`/`batch.xlsx` (same fields, one row per file — see
`_load_batch_rows` in sync.py; a sidecar always wins over a matching batch
row). Once the whisper job (``sync.py``) has produced a matching
transcript, the `llm:` and `prompt:` fields (each a single model/prompt
name or a list) are expanded into the CARTESIAN PRODUCT of every listed
prompt against every listed model, each run against a hosted KubeAI LLM. If
`stt:` requested multiple STT models, EVERY transcript variant
(`transcriptions/<stem>_<sttmodel>.txt`) is processed independently — the
result is uploaded to a sibling ``llm/`` folder as
``<transcript_stem>_<model>_<prompt>.md``, i.e. the STT model (if any) is
folded into the filename automatically since it's already part of
`<transcript_stem>`. An optional `context:` field is injected into prompt
templates via a `{context}` placeholder (same plain-string-replace
mechanism as `{transcript}`). Any OTHER field beyond the fixed
`stt`/`llm`/`prompt`/`context` schema (e.g. a `ground_truth` column added
to a batch.xlsx) is likewise available as a `{field_name}` placeholder —
see `_sidecar_from_dict`'s `extra_fields`.

Jobs are processed grouped by model (not crawl-discovery order) so that
consecutive calls to the same KubeAI-hosted model land within its
scaleDownDelaySeconds warm window, avoiding repeated cold starts.

For batch-file-driven folders, an optional `judges/<name>.yaml` per top-level
NEXTCLOUD_FOLDER entry (sibling to `prompts/`) scores every existing output
cell of its `jurisdiction:` prompt(s) — either `kind: label` (the judge LLM
states its score directly) or `kind: logprob` (G-Eval-style, score derived
from probability-weighted token logprobs). Results land in
`llm/<judge>_scores.xlsx` — one sheet per jurisdiction prompt plus an
"Overview" comparison sheet, so a judge's take across different prompts is
comparable in one file, not scattered across several. Already-scored cells
are never re-scored. Arbitrary extra sidecar/batch-row fields (e.g. a
user-added `ground_truth` column) are available as `{field_name}`
placeholders in both prompt and judge templates. See README_llm.md's
"Judges" section for the full schema.

Fully independent of the whisper job — no audio download, no diarization, no
Whisper calls. Files with neither a sidecar nor a batch row are never touched.

All configuration is read from environment variables (see README_llm.md).
"""

import asyncio
import json
import logging
import math
import os
import re
import shutil
import tempfile
from pathlib import Path

import aiohttp
import yaml
from dotenv import load_dotenv
from src.nextcloud.sync import (
    AUDIO_VIDEO_MIME_TYPES,
    SIDECAR_SUFFIXES,
    _download_text,
    _load_batch_rows,
    _make_webdav_client,
    _normalize_list,
    _stt_models_from_dict,
)
from tqdm import tqdm
from webdav3.client import Client
from webdav3.exceptions import RemoteResourceNotFound

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
)
logger = logging.getLogger(__name__)

TRANSCRIPT_SUBFOLDER = "transcriptions"
LLM_SUBFOLDER = "llm"
PROMPTS_SUBFOLDER = "prompts"
JUDGES_SUBFOLDER = "judges"
LOCAL_PROMPTS_DIR = Path(__file__).parent / "prompts"
DEFAULT_PROMPT = "summary"

# Max candidate tokens vLLM/OpenAI returns per generated-token position when
# logprobs are requested. No max_tokens cap is set on judge calls (see
# _call_judge_llm) — a reasoning-tuned judge model emits <think>...</think>
# as a mandatory part of its generation order, so truncating the token
# budget would cut it off mid-thought before it ever reaches the answer.
JUDGE_TOP_LOGPROBS = 20


# ---------------------------------------------------------------------------
# Sidecar / job discovery
# ---------------------------------------------------------------------------


_RESERVED_SIDECAR_FIELDS = frozenset({"stt", "llm", "prompt", "context"})


def _sidecar_from_dict(data: dict) -> dict:
    """Normalize an already-parsed sidecar/batch-row dict into model/prompt lists + context.

    Shared core of `_parse_sidecar` (YAML sidecar) and the batch-file path
    (`_load_batch_rows` in sync.py already produces this dict shape
    directly, no YAML round-trip needed).

    An empty dict (or one with fields omitted) falls back to a 1-element
    ``[LLM_DEFAULT_MODEL]`` / ``[DEFAULT_PROMPT]``. `stt` is intentionally
    ignored here — that field belongs to sync.py's whisper stage (see
    `_stt_models_from_dict`).

    Any field beyond the fixed schema (`stt`/`llm`/`prompt`/`context`) is
    passed through as `extra_fields` — lets a user add an arbitrary column
    (e.g. a `ground_truth` column in batch.xlsx, or an extra key in a YAML
    sidecar) and reference it as a `{field_name}` placeholder in both
    prompt templates (`_render_prompt`) and judge templates
    (`_render_judge_prompt`), same plain-substitution mechanism as
    `{context}`.

    Returns:
        Dict with keys "models" (list[str]), "prompts" (list[str]),
        "context" (str, empty if absent), and "extra_fields"
        (dict[str, str], empty if none present).
    """
    models = _normalize_list(data.get("llm")) or [os.environ.get("LLM_DEFAULT_MODEL", "glm-4-7-flash")]
    prompts = _normalize_list(data.get("prompt")) or [DEFAULT_PROMPT]
    context = str(data.get("context") or "")
    extra_fields = {
        str(k): str(v) for k, v in data.items() if k not in _RESERVED_SIDECAR_FIELDS and v is not None
    }
    return {"models": models, "prompts": prompts, "context": context, "extra_fields": extra_fields}


def _parse_sidecar(text: str) -> dict:
    """Parse a sidecar YAML's content into normalized model/prompt lists + context.

    Args:
        text: Raw YAML file contents.

    Returns:
        See `_sidecar_from_dict`.
    """
    return _sidecar_from_dict(yaml.safe_load(text) or {})


def _collect_llm_jobs(
    client: Client, remote_roots: list[str], tmp_dir: Path
) -> tuple[list[dict], dict[str, dict[str, dict]], dict[str, str]]:
    """Recursively find sidecar-driven LLM jobs ready to run.

    A job is ready when: a `<stem>.yaml`/`.yml` sidecar sits next to an
    audio/video file (or, absent that, the file has a matching row in this
    folder's batch.csv/batch.xlsx — see `_load_batch_rows` in sync.py; a
    sidecar always wins over a matching batch row), at least one
    `transcriptions/<stem>[_<sttmodel>].txt` already exists, and
    `llm/<transcript_stem>_<model>_<prompt>.md` does not exist yet. Every
    existing transcript variant for the stem is processed independently —
    `stt: [turbo, qwen3-asr-1.7b]` produces two full llm x prompt batches,
    one per STT model's transcript, each named after that transcript's own
    stem (so the STT model is folded into the output filename
    automatically). Each batch is the cartesian product of `llm:` models ×
    `prompt:` prompts (each field may be a single name or a list — see
    `_sidecar_from_dict`).

    Args:
        client: WebDAV client.
        remote_roots: Root remote folders to scan (see _make_webdav_client).
        tmp_dir: Local scratch directory for downloading sidecars.

    Returns:
        Tuple of (jobs, batch_folders, folder_roots):
        - jobs: list of dicts (transcript_path, llm_dir, output_name, model,
          prompt, context, remote_root — the folder this job's sidecar came
          from, needed so prompt-template lookups check the right folder's
          own `prompts/` subfolder first).
        - batch_folders: {folder_path: batch_rows} for every folder that had
          a batch.csv/batch.xlsx, used by `_build_prompt_tables` to build
          per-prompt review tables after processing.
        - folder_roots: {folder_path: top_level_root} for every folder in
          batch_folders — judges resolve once per top-level root (like
          prompts/), so the judge-scoring pass needs to know which root a
          nested batch folder belongs to.
    """
    jobs: list[dict] = []
    batch_folders: dict[str, dict[str, dict]] = {}
    folder_roots: dict[str, str] = {}

    def _walk(path: str, root: str) -> None:
        try:
            entries = client.list(path, get_info=True)
        except RemoteResourceNotFound:
            logger.warning("Remote path not found, skipping: %s", path)
            return

        subdirs = [
            e
            for e in entries
            if e["isdir"]
            and e["path"].rstrip("/") != path.rstrip("/")
            and Path(e["path"]).name
            not in (TRANSCRIPT_SUBFOLDER, LLM_SUBFOLDER, PROMPTS_SUBFOLDER, JUDGES_SUBFOLDER)
        ]
        files = [e for e in entries if not e["isdir"]]
        audio_stems = {
            Path(f["path"]).stem for f in files if f.get("content_type", "") in AUDIO_VIDEO_MIME_TYPES
        }
        sidecars_by_stem = {
            Path(f["path"]).stem: f["path"]
            for f in files
            if Path(f["path"]).suffix.lower() in SIDECAR_SUFFIXES and Path(f["path"]).stem in audio_stems
        }
        batch_rows = _load_batch_rows(client, files, path, tmp_dir)
        if batch_rows:
            batch_folders[path] = batch_rows
            folder_roots[path] = root
        relevant_stems = (set(sidecars_by_stem) | set(batch_rows)) & audio_stems

        if relevant_stems:
            transcript_dir = path.rstrip("/") + f"/{TRANSCRIPT_SUBFOLDER}/"
            try:
                # .txt only — every transcript has a matching .srt with the
                # same stem, and we'd otherwise see each transcript twice.
                existing_transcripts = {
                    Path(f).stem for f in client.list(transcript_dir) if f.lower().endswith(".txt")
                }
            except RemoteResourceNotFound:
                existing_transcripts = set()

            llm_dir = path.rstrip("/") + f"/{LLM_SUBFOLDER}/"
            try:
                existing_outputs = set(client.list(llm_dir))
            except RemoteResourceNotFound:
                existing_outputs = set()

            for stem in relevant_stems:
                sidecar_path = sidecars_by_stem.get(stem)
                if sidecar_path:
                    try:
                        text = _download_text(client, sidecar_path, tmp_dir)
                    except Exception as exc:
                        logger.error("Failed to download sidecar %s: %s", sidecar_path, exc)
                        continue
                    data = yaml.safe_load(text) or {}
                else:
                    # No per-file sidecar — fall back to this folder's batch.csv/batch.xlsx row.
                    data = batch_rows[stem]

                parsed = _sidecar_from_dict(data)

                # `stt` can produce several transcripts for one audio file:
                # the bare `<stem>.txt` (no `stt` — default, unsuffixed) or
                # `<stem>_<model>.txt` per listed STT model. Derive the
                # expected stems from the sidecar/batch-row data itself, not
                # by guessing from the folder listing — a naive `startswith`
                # match against every transcript stem in the directory is
                # ambiguous (e.g. "interview_2" is a different audio file's
                # transcript per the speaker-count filename convention in
                # sync.py, not an "interview" file with stt-suffix "2").
                stt_models = _stt_models_from_dict(data)
                expected_stems = [stem] if not stt_models else [f"{stem}_{m}" for m in stt_models]
                transcript_stems = [s for s in expected_stems if s in existing_transcripts]
                if not transcript_stems:
                    logger.info("Transcript not ready yet, will retry later: %s", stem)
                    continue
                if len(transcript_stems) < len(expected_stems):
                    logger.info(
                        "Some transcripts not ready yet for %s, processing what exists (%d/%d): %s",
                        stem, len(transcript_stems), len(expected_stems), transcript_stems,
                    )

                for transcript_stem in transcript_stems:
                    for model in parsed["models"]:
                        for prompt in parsed["prompts"]:
                            output_name = f"{transcript_stem}_{model}_{prompt}.md"
                            if output_name in existing_outputs:
                                continue

                            jobs.append(
                                {
                                    "transcript_path": transcript_dir + f"{transcript_stem}.txt",
                                    "llm_dir": llm_dir,
                                    "output_name": output_name,
                                    "model": model,
                                    "prompt": prompt,
                                    "context": parsed["context"],
                                    "extra_fields": parsed["extra_fields"],
                                    "remote_root": root,
                                }
                            )

        for d in subdirs:
            _walk(d["path"], root)

    for root in remote_roots:
        _walk(root, root)
    return jobs, batch_folders, folder_roots


def _scan_prompt_cells(
    client: Client, folder_path: str, batch_rows: dict[str, dict]
) -> dict[str, dict[str, dict[str, str]]]:
    """Map prompt -> transcript_stem -> model -> output filename, for outputs
    that actually exist in this batch folder's llm/ right now.

    Extracted from `_build_prompt_tables` (its original inline
    existing_transcripts/existing_outputs/expected_stems logic) since the
    judge-scoring pass needs the identical "what's actually been produced"
    map — both callers need to know which (stem, model, prompt) cells have a
    real `.md` output before doing anything with them.

    Args:
        client: WebDAV client.
        folder_path: Remote folder that has the batch.csv/batch.xlsx.
        batch_rows: {stem: row_dict}, as returned by `_load_batch_rows`.

    Returns:
        {prompt: {transcript_stem: {model: output_filename}}} — only
        entries whose output file actually exists in llm/.
    """
    transcript_dir = folder_path.rstrip("/") + f"/{TRANSCRIPT_SUBFOLDER}/"
    llm_dir = folder_path.rstrip("/") + f"/{LLM_SUBFOLDER}/"
    try:
        existing_transcripts = {Path(f).stem for f in client.list(transcript_dir) if f.lower().endswith(".txt")}
    except RemoteResourceNotFound:
        existing_transcripts = set()
    try:
        existing_outputs = set(client.list(llm_dir))
    except RemoteResourceNotFound:
        existing_outputs = set()

    cells: dict[str, dict[str, dict[str, str]]] = {}
    for stem, data in batch_rows.items():
        parsed = _sidecar_from_dict(data)
        stt_models = _stt_models_from_dict(data)
        expected_stems = [stem] if not stt_models else [f"{stem}_{m}" for m in stt_models]
        for transcript_stem in (s for s in expected_stems if s in existing_transcripts):
            for model in parsed["models"]:
                for prompt in parsed["prompts"]:
                    output_name = f"{transcript_stem}_{model}_{prompt}.md"
                    if output_name in existing_outputs:
                        cells.setdefault(prompt, {}).setdefault(transcript_stem, {})[model] = output_name
    return cells


def _build_prompt_tables(client: Client, folder_path: str, batch_rows: dict[str, dict], tmp_dir: Path) -> None:
    """Build one llm/<prompt>_table.xlsx per distinct prompt used in this
    folder's batch.csv/batch.xlsx.

    Batch-input usability feature: a customer who uploaded many files via a
    batch sheet gets one reviewable table per prompt instead of only
    scattered individual .md files (which still exist, unchanged) — rows
    are transcript stems (STT-model suffix included when `stt:` was a
    list), columns are LLM models, cells are that (stem, model, prompt)
    combination's output text (blank if not processed yet). Fully rebuilt
    from whatever's actually in llm/ right now (not incrementally appended)
    whenever called — the caller in `main()` only calls this for folders
    where a job actually uploaded something this run, so it's a consistent
    snapshot without needless rework on no-op runs. Only applies to
    batch-covered folders, not plain per-file YAML sidecars.

    Args:
        client: WebDAV client.
        folder_path: Remote folder that has the batch.csv/batch.xlsx.
        batch_rows: {stem: row_dict}, as returned by `_load_batch_rows`.
        tmp_dir: Local scratch directory for downloads/uploads.
    """
    llm_dir = folder_path.rstrip("/") + f"/{LLM_SUBFOLDER}/"
    cells = _scan_prompt_cells(client, folder_path, batch_rows)
    if not cells:
        return

    from openpyxl import Workbook

    for prompt, rows in cells.items():
        models = sorted({model for row in rows.values() for model in row})
        wb = Workbook()
        ws = wb.active
        ws.append(["filename"] + models)
        for transcript_stem in sorted(rows):
            row_values = [transcript_stem]
            for model in models:
                output_name = rows[transcript_stem].get(model)
                text = ""
                if output_name:
                    try:
                        text = _download_text(client, llm_dir + output_name, tmp_dir)
                    except Exception as exc:
                        logger.error("Failed to download %s for %s_table.xlsx: %s", output_name, prompt, exc)
                row_values.append(text)
            ws.append(row_values)

        local_path = tmp_dir / f"{prompt}_table.xlsx"
        wb.save(local_path)
        remote_out = llm_dir + f"{prompt}_table.xlsx"
        try:
            client.upload_sync(local_path=str(local_path), remote_path=remote_out)
            logger.info("Uploaded %s", remote_out)
        except Exception as exc:
            logger.error("Upload failed for %s: %s", remote_out, exc)
        local_path.unlink(missing_ok=True)


# ---------------------------------------------------------------------------
# Prompting / LLM call
# ---------------------------------------------------------------------------


def _load_prompt_template(
    client: Client,
    remote_root: str,
    prompt_name: str,
    tmp_dir: Path,
    cache: dict[str, str],
) -> str:
    """Load a prompt template, preferring the Nextcloud ``prompts/`` folder.

    Templates are edited by uploading a `.md` file to
    ``<remote_root>/prompts/<prompt_name>.md`` — no image rebuild or redeploy
    needed. Falls back to the template bundled in the image (under the
    script's local ``prompts/`` folder) only if the remote one is missing,
    so a fresh deployment still has a working default. Results are cached
    per run since many jobs typically share the same prompt.

    Args:
        client: WebDAV client.
        remote_root: Root remote folder this job's sidecar came from (prompts
            live at `<remote_root>/prompts/`) — with multiple NEXTCLOUD_FOLDER
            entries, each folder's own prompts/ is checked independently, so
            the cache key must include it (see `cache`).
        prompt_name: Stem of a `.md` template file.
        tmp_dir: Local scratch directory for the temporary download.
        cache: Dict reused across calls within a single run, keyed
            `"<remote_root>:<prompt_name>"` — the same prompt name can
            resolve to a different template per folder.

    Returns:
        Raw template text (not yet rendered with a transcript).

    Raises:
        FileNotFoundError: If no matching template exists remotely or locally.
    """
    cache_key = f"{remote_root}:{prompt_name}"
    if cache_key in cache:
        return cache[cache_key]

    remote_path = f"{remote_root.rstrip('/')}/{PROMPTS_SUBFOLDER}/{prompt_name}.md"
    try:
        template = _download_text(client, remote_path, tmp_dir)
        logger.info("Loaded prompt '%s' from Nextcloud (%s).", prompt_name, remote_root)
    except RemoteResourceNotFound:
        local_path = LOCAL_PROMPTS_DIR / f"{prompt_name}.md"
        if not local_path.exists():
            raise FileNotFoundError(
                f"Prompt '{prompt_name}' not found in Nextcloud ({remote_path}) "
                f"or bundled defaults ({local_path})."
            )
        logger.info("Prompt '%s' not in Nextcloud, using bundled default.", prompt_name)
        template = local_path.read_text(encoding="utf-8")

    cache[cache_key] = template
    return template


def _render_prompt(template: str, transcript: str, context: str = "", extra_fields: dict[str, str] | None = None) -> str:
    """Render a prompt template with the transcript text and optional context/extra fields.

    Args:
        template: Raw template text (from `_load_prompt_template`).
        transcript: Plain-text transcript content.
        context: Sidecar-provided `context:` text, injected via a `{context}`
            placeholder (same plain-string-replace mechanism as
            `{transcript}`). Templates that don't reference `{context}` are
            unaffected — empty string is the default, matching every sidecar
            that omits the field.
        extra_fields: Any non-schema sidecar/batch-row fields (see
            `_sidecar_from_dict`), each substituted as its own
            `{field_name}` placeholder. A template that doesn't reference a
            given field name is unaffected.

    Returns:
        Rendered prompt text.
    """
    rendered = template.replace("{transcript}", transcript).replace("{context}", context)
    for key, value in (extra_fields or {}).items():
        rendered = rendered.replace("{" + key + "}", value)
    return rendered


async def _call_llm(session: aiohttp.ClientSession, model: str, prompt: str) -> str | None:
    """Send a chat-completion request to the in-cluster KubeAI endpoint.

    Args:
        session: Shared aiohttp session.
        model: KubeAI model name.
        prompt: Fully rendered prompt text.

    Returns:
        The assistant's reply text, or None on failure.
    """
    llm_url = os.environ.get("LLM_URL", "http://kubeai.llm.svc.cluster.local/openai/v1")
    endpoint = f"{llm_url.rstrip('/')}/chat/completions"
    payload = {"model": model, "messages": [{"role": "user", "content": prompt}]}
    headers = {"Authorization": "Bearer not-used"}

    try:
        async with session.post(endpoint, json=payload, headers=headers) as resp:
            if not (200 <= resp.status < 400):
                body = await resp.text()
                logger.error("LLM request failed (HTTP %s): %s", resp.status, body)
                return None
            data = json.loads(await resp.read())
            content = data["choices"][0]["message"]["content"]
            # Reasoning models (e.g. glm-4-7-flash) can emit their chain-of-thought
            # inline as `<think>...</think>` instead of a separate "reasoning"
            # field, depending on vLLM's reasoning-parser config. Strip it.
            if "</think>" in content:
                content = content.rsplit("</think>", 1)[1]
            return content.strip()
    except Exception as exc:
        # asyncio.TimeoutError (and a few other exceptions) have an empty
        # str() — include the type name so timeouts are distinguishable
        # from other failures in the logs.
        logger.error("Error calling LLM: %s: %s", type(exc).__name__, exc)
        return None


# ---------------------------------------------------------------------------
# Judges (LLM-as-judge)
# ---------------------------------------------------------------------------


def _parse_judge_scale(scale) -> dict:
    """Parse a judge YAML's `scale:` field into a normalized shape.

    Accepts:
    - `"1-5"` (a string, `low-high`) or a 2-element numeric list `[1, 5]` ->
      an inclusive numeric range.
    - A >2-element numeric list, e.g. `[1, 3, 5]` -> an explicit discrete
      numeric set (supports non-contiguous scales).
    - A list of non-numeric strings, e.g. `["poor", "fair", "good"]` -> a
      label scale. Must be given in ASCENDING QUALITY ORDER — required for
      `kind: logprob`'s ordinal mapping (see `_score_from_logprobs`).

    Args:
        scale: The raw `scale:` value from a parsed judge YAML.

    Returns:
        `{"type": "numeric"|"label", "values": list}`.

    Raises:
        ValueError: If `scale` doesn't match any of the above shapes.
    """
    if isinstance(scale, str):
        m = re.match(r"^\s*(-?\d+)\s*-\s*(-?\d+)\s*$", scale)
        if m:
            low, high = int(m.group(1)), int(m.group(2))
            return {"type": "numeric", "values": list(range(low, high + 1))}
    elif isinstance(scale, list) and scale:
        try:
            nums = [int(v) if float(v).is_integer() else float(v) for v in scale]
        except (TypeError, ValueError):
            nums = None
        if nums is not None:
            if len(nums) == 2:
                low, high = nums
                if isinstance(low, int) and isinstance(high, int):
                    return {"type": "numeric", "values": list(range(low, high + 1))}
            return {"type": "numeric", "values": nums}
        return {"type": "label", "values": [str(v) for v in scale]}

    raise ValueError(f"Judge scale must be a numeric range/list or a list of labels, got: {scale!r}")


def _normalize_anchors(anchors, scale: dict) -> dict[str, str]:
    """Normalize a judge YAML's `anchors:` field into `{scale_value_str: description}`.

    A user should never be forced to write one anchor per scale point,
    especially on a fine-grained numeric scale (e.g. 1-100). Two input
    shapes are accepted:
    - dict (explicit): `{scale_value: description}` — any sparse subset,
      used as-is at exactly those positions.
    - list (auto-distributed): plain list of description strings with no
      explicit positions — spread evenly across `scale["values"]`,
      endpoints inclusive. For N anchors over M scale values, anchor `i`
      lands at scale index `round(i * (M-1) / (N-1))` (N==1 -> the middle
      scale value).

    Args:
        anchors: The raw `anchors:` value (dict, list, or None/absent).
        scale: Parsed scale, as returned by `_parse_judge_scale`.

    Returns:
        `{str(scale_value): description}`, possibly empty.
    """
    if not anchors:
        return {}
    if isinstance(anchors, dict):
        return {str(k): str(v) for k, v in anchors.items()}

    values = scale["values"]
    n, m = len(anchors), len(values)
    if n == 1:
        positions = [m // 2]
    else:
        positions = [round(i * (m - 1) / (n - 1)) for i in range(n)]
    return {str(values[pos]): str(desc) for pos, desc in zip(positions, anchors)}


def _format_scale_section(scale: dict) -> str:
    """Render the scale portion of a judge prompt."""
    if scale["type"] == "numeric":
        return f"Scale: rate from {scale['values'][0]} to {scale['values'][-1]}."
    return f"Scale: choose one of: {', '.join(scale['values'])}."


def _format_anchors_section(anchors, scale: dict) -> str:
    """Render the anchors portion of a judge prompt, or "" if none given."""
    normalized = _normalize_anchors(anchors, scale)
    if not normalized:
        return ""
    lines = "\n".join(f"- {value}: {desc}" for value, desc in normalized.items())
    return f"Anchors:\n{lines}"


def _format_few_shot_section(few_shot) -> str:
    """Render the few-shot examples portion of a judge prompt, or "" if none given."""
    if not few_shot:
        return ""
    blocks = []
    for ex in few_shot:
        block = f"Response: {ex.get('response', '')}\nScore: {ex.get('score', '')}"
        if ex.get("rationale"):
            block += f"\nRationale: {ex['rationale']}"
        blocks.append(block)
    return "Examples:\n" + "\n\n".join(blocks)


def _format_anti_pattern_section(anti_pattern) -> str:
    """Render the anti-pattern portion of a judge prompt, or "" if none given."""
    items = _normalize_list(anti_pattern)
    if not items:
        return ""
    lines = "\n".join(f"- {item}" for item in items)
    return f"Do NOT reward:\n{lines}"


def _build_judge_template(judge_data: dict) -> str:
    """Assemble one judge's fixed-order prompt template from its parsed YAML.

    Fixed section order (only non-empty/optional ones are included):
      1. `judge_data["prompt"]` (stripped) — the user's own rubric text.
      2. Scale section — always present.
      3. Anchors section — only if `anchors` given.
      4. Few-shot examples section — only if `few_shot` given.
      5. Anti-pattern section — only if `anti_pattern` given.
      6. "Response to evaluate:\\n{response}" — always present.
      7. A kind-specific answer-format instruction — always present.

    Sections are joined with blank lines. Returns the unrendered template
    (still contains `{response}`, and `{transcript}`/`{context}` if section
    1 referenced them) — built once per judge at discovery time, not per
    scored cell.

    Args:
        judge_data: The raw parsed judge YAML dict (already validated).

    Returns:
        Unrendered judge prompt template.
    """
    scale = judge_data["_scale"]
    sections = [
        judge_data["prompt"].strip(),
        _format_scale_section(scale),
        _format_anchors_section(judge_data.get("anchors"), scale),
        _format_few_shot_section(judge_data.get("few_shot")),
        _format_anti_pattern_section(judge_data.get("anti_pattern")),
        "Response to evaluate:\n{response}",
        (
            "Respond with ONLY the label above, nothing else. No explanation, no punctuation."
            if judge_data["kind"] == "label"
            else "Respond with a single token: the number only. Do not explain your reasoning."
        ),
    ]
    return "\n\n".join(s for s in sections if s)


def _render_judge_prompt(
    template: str,
    response: str,
    transcript: str = "",
    context: str = "",
    extra_fields: dict[str, str] | None = None,
) -> str:
    """Render a judge prompt template with the response text being judged.

    Args:
        template: Unrendered template from `_build_judge_template`.
        response: The `.md` output text being judged.
        transcript: Original transcript text, substituted only if the
            judge's own text references `{transcript}`.
        context: Sidecar-provided context, substituted only if referenced.
        extra_fields: Any non-schema sidecar/batch-row fields from the
            judged file's own row (see `_sidecar_from_dict`) — e.g. a
            user-added `ground_truth` column, referenced as `{ground_truth}`
            in the judge's `prompt:` text for real reference-based judging,
            not just reference-free quality checks.

    Returns:
        Rendered judge prompt text.
    """
    rendered = (
        template.replace("{response}", response).replace("{transcript}", transcript).replace("{context}", context)
    )
    for key, value in (extra_fields or {}).items():
        rendered = rendered.replace("{" + key + "}", value)
    return rendered


def _parse_judge_yaml(text: str, name: str) -> dict:
    """Parse and validate one `judges/<name>.yaml`.

    Args:
        text: Raw YAML file contents.
        name: Judge name (the file's stem), used in error messages and as
            the output filename prefix.

    Returns:
        `{"name", "llm", "kind", "jurisdiction", "scale", "template"}`.

    Raises:
        ValueError: If a required field is missing, `kind` is invalid, or
            the scale doesn't parse (see `_parse_judge_scale`). Callers
            should catch this per-file so one malformed judge doesn't abort
            discovery of every other judge.
    """
    data = yaml.safe_load(text) or {}
    for field in ("prompt", "jurisdiction", "scale", "llm", "kind"):
        if not data.get(field):
            raise ValueError(f"judges/{name}.yaml: missing required field '{field}'")
    if data["kind"] not in ("label", "logprob"):
        raise ValueError(f"judges/{name}.yaml: kind must be 'label' or 'logprob', got {data['kind']!r}")

    scale = _parse_judge_scale(data["scale"])
    data["_scale"] = scale  # stashed for _build_judge_template, not part of the public return shape
    return {
        "name": name,
        "llm": str(data["llm"]),
        "kind": data["kind"],
        "jurisdiction": _normalize_list(data["jurisdiction"]),
        "scale": scale,
        "template": _build_judge_template(data),
    }


def _collect_judges(client: Client, remote_roots: list[str], tmp_dir: Path) -> dict[str, list[dict]]:
    """List `<root>/judges/*.{yaml,yml}` for each remote_root.

    Non-recursive — `judges/` is a single per-root folder, same convention
    as `prompts/` (see `_load_prompt_template`'s docstring: `remote_root` is
    always a top-level configured folder, never a nested one).

    Args:
        client: WebDAV client.
        remote_roots: Root remote folders to scan.
        tmp_dir: Local scratch directory for downloading judge YAMLs.

    Returns:
        `{remote_root: [judge_config, ...]}`. A malformed judge YAML logs an
        error and is skipped — it does not abort discovery of other judges.
    """
    result: dict[str, list[dict]] = {}
    for root in remote_roots:
        judges_dir = root.rstrip("/") + f"/{JUDGES_SUBFOLDER}/"
        try:
            entries = client.list(judges_dir, get_info=True)
        except RemoteResourceNotFound:
            continue
        judges = []
        for e in entries:
            if e["isdir"] or Path(e["path"]).suffix.lower() not in SIDECAR_SUFFIXES:
                continue
            name = Path(e["path"]).stem
            try:
                text = _download_text(client, e["path"], tmp_dir)
                judges.append(_parse_judge_yaml(text, name))
            except Exception as exc:
                logger.error("Failed to load judge '%s' in %s: %s", name, judges_dir, exc)
        if judges:
            result[root] = judges
    return result


def _locate_score_token(logprobs_content: list[dict], scale: dict) -> list[dict] | None:
    """Find the `top_logprobs` candidate list at the position holding the
    judge's answer token.

    Reasoning-tuned judge models (e.g. glm-4-7-flash without a configured
    vLLM `--reasoning-parser`) emit `<think>...</think>` inline as literal
    text in `content` — so the answer token is not necessarily the first
    generated token. This walks the token sequence, and once `</think>`
    appears in the concatenated running text, resumes scanning from the
    entry right after the one that completed it (if `</think>` never
    appears at all, scans from index 0). Confirmed live (2026-08-28,
    glm-4-7-flash on vLLM v0.28.0 with `--reasoning-parser=glm45`
    configured): even once a reasoning-parser splits the message-level
    `content`/`reasoning` fields cleanly, `logprobs.content` stays ONE flat
    raw token stream covering reasoning + answer together and still
    contains a literal `</think>` token — so this scan is needed
    unconditionally, not just as a fallback for the no-parser case. From the
    start index, returns the first entry whose own *sampled* token matches a
    valid scale value.

    Args:
        logprobs_content: `choices[0]["logprobs"]["content"]` from a judge
            chat-completion response — a list of
            `{"token", "logprob", "top_logprobs": [{"token","logprob"}, ...]}`.
        scale: Parsed scale, as returned by `_parse_judge_scale`.

    Returns:
        The matching entry's `top_logprobs` list, or `None` if no token in
        the sequence matches any valid scale value.
    """
    running, start = "", 0
    for i, entry in enumerate(logprobs_content):
        running += entry.get("token", "")
        if "</think>" in running:
            start = i + 1
            break

    valid = {str(v).lower() for v in scale["values"]}
    for entry in logprobs_content[start:]:
        if entry.get("token", "").strip().lower() in valid:
            return entry.get("top_logprobs") or None
    return None


def _score_from_logprobs(top_logprobs: list[dict], scale: dict) -> float | None:
    """Probability-weighted G-Eval score over the candidate tokens at one position.

    For a numeric scale, the candidate value is the number itself. For a
    label scale, the candidate value is the 0-based ordinal index in
    `scale["values"]` (ascending-quality order) — this is what makes
    `kind: logprob` produce a continuous score even for a discrete label
    scale.

    `weight(v) = exp(logprob)` for each `top_logprobs` entry whose token
    (stripped, case-insensitive) matches candidate value `v` (summed if the
    same value appears via more than one token variant, e.g. `"4"` and
    `" 4"`). `score = sum(v * weight(v)) / sum(weight(v))`, over matched `v`
    only.

    Args:
        top_logprobs: The candidate list from `_locate_score_token`.
        scale: Parsed scale, as returned by `_parse_judge_scale`.

    Returns:
        The weighted-average score, or `None` if zero entries match any
        scale value.
    """
    if scale["type"] == "numeric":
        candidates = {str(v).lower(): float(v) for v in scale["values"]}
    else:
        candidates = {v.lower(): float(i) for i, v in enumerate(scale["values"])}

    weights: dict[float, float] = {}
    for entry in top_logprobs:
        tok = str(entry.get("token", "")).strip().lower()
        if tok in candidates:
            v = candidates[tok]
            weights[v] = weights.get(v, 0.0) + math.exp(entry["logprob"])

    total = sum(weights.values())
    if total == 0:
        return None
    return sum(v * w for v, w in weights.items()) / total


def _score_from_label(text: str, scale: dict) -> float | str | None:
    """Parse a plain-text judge reply against the scale (`kind: label`).

    Numeric: strips whitespace/trailing punctuation and tries `int()`/
    `float()`; on failure, regex-searches for the first standalone integer
    substring that's a valid scale value. Label: exact match (stripped,
    case-insensitive) against `scale["values"]`; if no exact match, falls
    back to "exactly one label appears as a substring of the reply".

    Args:
        text: The judge LLM's raw reply text.
        scale: Parsed scale, as returned by `_parse_judge_scale`.

    Returns:
        The matched value (float for numeric, the original-cased label
        string for label scales), or `None` if parsing failed.
    """
    cleaned = text.strip().rstrip(".!")
    if scale["type"] == "numeric":
        try:
            return float(cleaned)
        except ValueError:
            pass
        valid = {str(v) for v in scale["values"]}
        for m in re.finditer(r"-?\d+(?:\.\d+)?", cleaned):
            if m.group(0) in valid:
                return float(m.group(0))
        return None

    for value in scale["values"]:
        if cleaned.lower() == value.lower():
            return value
    matches = [v for v in scale["values"] if v.lower() in cleaned.lower()]
    return matches[0] if len(matches) == 1 else None


async def _call_judge_llm(
    session: aiohttp.ClientSession, model: str, prompt: str, want_logprobs: bool
) -> tuple[str | None, list[dict] | None]:
    """Send a judge chat-completion request to the in-cluster KubeAI endpoint.

    Same endpoint/headers/error-handling/`</think>`-stripping shape as
    `_call_llm` (duplicated deliberately — `_call_llm` must not change for
    the non-judge path). No `max_tokens` cap is set (see `JUDGE_TOP_LOGPROBS`'s
    docstring note above: a reasoning-tuned model must be allowed to finish
    its `<think>` block before it ever reaches the answer).

    Args:
        session: Shared aiohttp session.
        model: KubeAI model name to use as judge.
        prompt: Fully rendered judge prompt text.
        want_logprobs: If True, requests `logprobs`/`top_logprobs` in the
            payload (for `kind: logprob`).

    Returns:
        `(content, logprobs_content)`:
        - content: stripped reply text, or `None` on any HTTP/network
          failure (identical contract to `_call_llm`'s return).
        - logprobs_content: the raw `choices[0]["logprobs"]["content"]` list
          when `want_logprobs` was True AND the server returned a
          well-formed non-empty logprobs field, else `None`. A `None` here
          (not an empty list, not an exception) is the signal that the
          server did not give us usable logprobs — the graceful-degradation
          contract for this cluster's unverified KubeAI/vLLM logprobs
          support.
    """
    llm_url = os.environ.get("LLM_URL", "http://kubeai.llm.svc.cluster.local/openai/v1")
    endpoint = f"{llm_url.rstrip('/')}/chat/completions"
    payload = {"model": model, "messages": [{"role": "user", "content": prompt}]}
    if want_logprobs:
        payload["logprobs"] = True
        payload["top_logprobs"] = JUDGE_TOP_LOGPROBS
    headers = {"Authorization": "Bearer not-used"}

    try:
        async with session.post(endpoint, json=payload, headers=headers) as resp:
            if not (200 <= resp.status < 400):
                body = await resp.text()
                logger.error("Judge LLM request failed (HTTP %s): %s", resp.status, body)
                return None, None
            data = json.loads(await resp.read())
            content = data["choices"][0]["message"]["content"]
            if "</think>" in content:
                content = content.rsplit("</think>", 1)[1]
            content = content.strip()

            logprobs_content = None
            if want_logprobs:
                try:
                    logprobs_content = data["choices"][0]["logprobs"]["content"] or None
                except (KeyError, TypeError):
                    logprobs_content = None
            return content, logprobs_content
    except Exception as exc:
        logger.error("Error calling judge LLM: %s: %s", type(exc).__name__, exc)
        return None, None


def _read_judge_scores_xlsx(local_path: Path) -> dict[str, dict[str, dict[str, float]]]:
    """Parse a downloaded `<judge>_scores.xlsx` into `{prompt: {stem: {model: score}}}`.

    Each jurisdiction prompt gets its own sheet, named after the prompt
    (truncated to Excel's 31-char sheet-name limit, and subject to
    openpyxl's own sanitizing/dedup of invalid characters — a strong
    assumption that jurisdiction prompt names are short simple slugs, true
    of every prompt name in this codebase's actual use so far). The
    "Overview" sheet (a derived cross-prompt comparison, see
    `_write_judge_scores_xlsx`) is skipped — it's not raw per-cell data to
    resume scoring from. Blank cells (unscored) are skipped, never treated
    as a score of 0/None.

    Args:
        local_path: Local path of a downloaded scores workbook.

    Returns:
        `{prompt: {stem: {model: score}}}`.
    """
    from openpyxl import load_workbook

    wb = load_workbook(local_path, read_only=True, data_only=True)
    result: dict[str, dict[str, dict[str, float]]] = {}
    for sheet_name in wb.sheetnames:
        if sheet_name == "Overview":
            continue
        rows = wb[sheet_name].iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            continue
        models = list(header[1:])

        prompt_scores: dict[str, dict[str, float]] = {}
        for row in rows:
            if not row or row[0] is None:
                continue
            stem = str(row[0])
            for model, value in zip(models, row[1:]):
                if value is not None:
                    prompt_scores.setdefault(stem, {})[str(model)] = float(value)
        if prompt_scores:
            result[sheet_name] = prompt_scores
    return result


def _load_existing_scores(client: Client, remote_path: str, tmp_dir: Path) -> dict[str, dict[str, dict[str, float]]]:
    """Download+parse an existing scores workbook, or `{}` if it doesn't exist yet.

    Args:
        client: WebDAV client.
        remote_path: Remote path of the `<judge>_scores.xlsx`.
        tmp_dir: Local scratch directory for the temporary download.

    Returns:
        `{prompt: {stem: {model: score}}}`, `{}` if the file doesn't exist remotely.
    """
    local_path = tmp_dir / Path(remote_path).name
    try:
        client.download_sync(remote_path=remote_path, local_path=str(local_path))
    except RemoteResourceNotFound:
        return {}
    try:
        return _read_judge_scores_xlsx(local_path)
    finally:
        local_path.unlink(missing_ok=True)


def _batch_row_for_transcript_stem(transcript_stem: str, batch_rows: dict[str, dict]) -> dict | None:
    """Find the batch row whose file produced a given transcript stem.

    `batch_rows` is keyed by the original audio file's stem, but a
    transcript stem can be `<stem>` (no `stt:` override) or `<stem>_<model>`
    (one per listed STT model) — the same expected-stems derivation
    `_scan_prompt_cells`/`_collect_llm_jobs` use, needed here to find the
    original row's `context:` field for a given already-scanned cell.

    Args:
        transcript_stem: A transcript stem as it appears in `transcriptions/`.
        batch_rows: `{stem: row_dict}`, as returned by `_load_batch_rows`.

    Returns:
        The matching row dict, or `None` if no row produces this stem.
    """
    for stem, data in batch_rows.items():
        stt_models = _stt_models_from_dict(data)
        expected = [stem] if not stt_models else [f"{stem}_{m}" for m in stt_models]
        if transcript_stem in expected:
            return data
    return None


def _missing_judge_cells(
    cells: dict[str, dict[str, str]], existing_scores: dict[str, dict[str, float]]
) -> list[tuple[str, str, str]]:
    """Pure set-difference: cells needing a score.

    This is the entire "don't re-score" guarantee — isolated as its own
    pure, I/O-free function specifically so it can be unit-tested in
    isolation.

    Args:
        cells: `{stem: {model: output_filename}}` — one prompt's worth, from
            `_scan_prompt_cells`.
        existing_scores: `{stem: {model: score}}`, from `_load_existing_scores`.

    Returns:
        List of `(stem, model, output_filename)` triples present in `cells`
        but absent from `existing_scores`.
    """
    return [
        (stem, model, filename)
        for stem, models in cells.items()
        for model, filename in models.items()
        if model not in existing_scores.get(stem, {})
    ]


def _collect_judge_scoring_groups(
    client: Client,
    batch_folders: dict[str, dict[str, dict]],
    folder_roots: dict[str, str],
    judges_by_root: dict[str, list[dict]],
    tmp_dir: Path,
) -> list[dict]:
    """Build one scoring group per (batch folder, judge) with at least one
    jurisdiction prompt having existing LLM output cells.

    One group == one output workbook (`llm/<judge>_scores.xlsx`) — a judge
    scoring multiple jurisdiction prompts in the same folder gets ONE file
    with one sheet per prompt, not one file per prompt, so a person can
    compare a judge's take across different prompts side by side instead
    of hunting across scattered files.

    Args:
        client: WebDAV client.
        batch_folders: `{folder_path: batch_rows}`, from `_collect_llm_jobs`.
        folder_roots: `{folder_path: top_level_root}`, from `_collect_llm_jobs`.
        judges_by_root: `{remote_root: [judge_config, ...]}`, from `_collect_judges`.
        tmp_dir: Local scratch directory for downloads.

    Returns:
        List of dicts: `{folder_path, llm_dir, judge, batch_rows,
        scores_remote_path, prompts: {prompt: {cells, existing_scores
        (loaded once, mutated in place by the caller as new scores land)}}}`.
    """
    groups = []
    for folder_path, batch_rows in batch_folders.items():
        judges = judges_by_root.get(folder_roots[folder_path], [])
        if not judges:
            continue
        prompt_cells = _scan_prompt_cells(client, folder_path, batch_rows)
        llm_dir = folder_path.rstrip("/") + f"/{LLM_SUBFOLDER}/"
        for judge in judges:
            scores_remote_path = llm_dir + f"{judge['name']}_scores.xlsx"
            existing_by_prompt = _load_existing_scores(client, scores_remote_path, tmp_dir)
            prompts_data: dict[str, dict] = {}
            for prompt in judge["jurisdiction"]:
                cells = prompt_cells.get(prompt)
                if not cells:
                    continue
                prompts_data[prompt] = {
                    "cells": cells,
                    "existing_scores": existing_by_prompt.get(prompt, {}),
                }
            if not prompts_data:
                continue
            groups.append(
                {
                    "folder_path": folder_path,
                    "llm_dir": llm_dir,
                    "judge": judge,
                    "batch_rows": batch_rows,
                    "prompts": prompts_data,
                    "scores_remote_path": scores_remote_path,
                }
            )
    return groups


def _write_judge_scores_xlsx(local_path: Path, prompts_data: dict[str, dict[str, dict[str, float]]]) -> None:
    """Write one judge's scores workbook: one sheet per jurisdiction prompt,
    plus an "Overview" sheet comparing prompts against each other.

    Per-prompt sheet (named after the prompt, truncated to Excel's 31-char
    limit): header `["filename"] + models`, one row per stem (sorted),
    numeric cell per (stem, model), blank (`None`, not 0) if unscored —
    blank is what makes `_missing_judge_cells` correctly treat it as
    still-needs-scoring on a future run. A bar chart is embedded on the
    same sheet, built directly off that sheet's own cell range.

    "Overview" (first sheet): one row per prompt, one column per model,
    cell = that (prompt, model)'s average score across all its stems — the
    actual "compare prompts in this judge's jurisdiction" view, since the
    per-prompt sheets alone only let you compare models within one prompt.
    Plus its own bar chart (categories = prompts, one series per model).

    Full rewrite of the file's contents each call (old + new scores already
    merged into each prompt's score map by the caller before this is
    invoked).

    Args:
        local_path: Local path to save the workbook to.
        prompts_data: `{prompt: {stem: {model: score}}}` — every
            jurisdiction prompt this judge has data for in this folder.
    """
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference

    wb = Workbook()
    wb.remove(wb.active)

    overview_rows: list[tuple[str, dict[str, float]]] = []
    for prompt in sorted(prompts_data):
        score_map = prompts_data[prompt]
        models = sorted({m for row in score_map.values() for m in row})
        ws = wb.create_sheet(prompt[:31])
        ws.append(["filename"] + models)
        for stem in sorted(score_map):
            ws.append([stem] + [score_map[stem].get(m) for m in models])

        n_rows = len(score_map)
        if models and n_rows:
            chart = BarChart()
            chart.type, chart.title = "col", prompt
            chart.y_axis.title, chart.x_axis.title = "score", "transcript"
            data = Reference(ws, min_col=2, max_col=1 + len(models), min_row=1, max_row=1 + n_rows)
            cats = Reference(ws, min_col=1, min_row=2, max_row=1 + n_rows)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            ws.add_chart(chart, ws.cell(row=1, column=len(models) + 3).coordinate)

        averages = {
            model: sum(row[model] for row in score_map.values() if model in row)
            / sum(1 for row in score_map.values() if model in row)
            for model in models
            if any(model in row for row in score_map.values())
        }
        overview_rows.append((prompt, averages))

    overview = wb.create_sheet("Overview", 0)
    all_models = sorted({m for _, avg in overview_rows for m in avg})
    overview.append(["prompt"] + all_models)
    for prompt, avg in overview_rows:
        overview.append([prompt] + [avg.get(m) for m in all_models])

    n_prompts = len(overview_rows)
    if all_models and n_prompts:
        chart = BarChart()
        chart.type, chart.title = "col", "Prompt comparison"
        chart.y_axis.title, chart.x_axis.title = "avg score", "prompt"
        data = Reference(overview, min_col=2, max_col=1 + len(all_models), min_row=1, max_row=1 + n_prompts)
        cats = Reference(overview, min_col=1, min_row=2, max_row=1 + n_prompts)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        overview.add_chart(chart, overview.cell(row=1, column=len(all_models) + 3).coordinate)

    wb.save(local_path)


def _upload_judge_group_scores(client: Client, group: dict, tmp_dir: Path) -> None:
    """Write and upload one group's scores.xlsx from its current `existing_scores`.

    Called as soon as a group's cells are all attempted (see `main()`) — not
    batched to the end of the whole judge pass, so a large backlog's results
    become visible incrementally, group by group, instead of only after
    every single cell across every judge has been scored.

    Args:
        client: WebDAV client.
        group: One entry from `_collect_judge_scoring_groups`'s result, with
            each prompt's `existing_scores` already updated with this run's
            new scores.
        tmp_dir: Local scratch directory for the upload.
    """
    prompts_data = {prompt: pdata["existing_scores"] for prompt, pdata in group["prompts"].items()}
    local = tmp_dir / Path(group["scores_remote_path"]).name
    _write_judge_scores_xlsx(local, prompts_data)
    try:
        client.list(group["llm_dir"])
    except RemoteResourceNotFound:
        client.mkdir(group["llm_dir"])
    try:
        client.upload_sync(local_path=str(local), remote_path=group["scores_remote_path"])
        logger.info("Uploaded %s", group["scores_remote_path"])
    except Exception as exc:
        logger.error("Upload failed for %s: %s", group["scores_remote_path"], exc)
    local.unlink(missing_ok=True)


async def _score_one_judge_cell(
    client: Client, session: aiohttp.ClientSession, cell_job: dict, tmp_dir: Path
) -> float | str | None:
    """Score one (stem, model) cell against one judge.

    Downloads the output `.md` being judged, downloads the transcript only
    if the judge's template references `{transcript}` (avoids a needless
    WebDAV round-trip otherwise), renders the judge prompt, and dispatches
    to label or logprob scoring per `judge["kind"]`. Any exception is
    caught, logged, and returns `None` — never propagates (matches
    `_call_llm`'s and the main jobs loop's per-item try/except discipline).

    Args:
        client: WebDAV client.
        session: Shared aiohttp session.
        cell_job: `{**group, "stem", "model", "output_name"}` — one flattened
            work item from `_collect_judge_scoring_groups`'s groups.
        tmp_dir: Local scratch directory for downloads.

    Returns:
        The score (float, or a label string — though `kind: label` on a
        label scale returns the original-cased label, which the caller
        stores as-is), or `None` if scoring failed/was skipped.
    """
    judge = cell_job["judge"]
    try:
        response_text = _download_text(client, cell_job["llm_dir"] + cell_job["output_name"], tmp_dir)
    except Exception as exc:
        logger.error("Judge '%s': failed to download %s: %s", judge["name"], cell_job["output_name"], exc)
        return None

    transcript = ""
    if "{transcript}" in judge["template"]:
        transcript_path = cell_job["folder_path"].rstrip("/") + f"/{TRANSCRIPT_SUBFOLDER}/{cell_job['stem']}.txt"
        try:
            transcript = _download_text(client, transcript_path, tmp_dir)
        except Exception as exc:
            logger.error("Judge '%s': failed to download transcript for %s: %s", judge["name"], cell_job["stem"], exc)

    context = ""
    extra_fields: dict[str, str] = {}
    row_data = _batch_row_for_transcript_stem(cell_job["stem"], cell_job["batch_rows"])
    if row_data is not None:
        parsed_row = _sidecar_from_dict(row_data)
        context = parsed_row["context"]
        extra_fields = parsed_row["extra_fields"]

    prompt = _render_judge_prompt(judge["template"], response_text, transcript, context, extra_fields)

    if judge["kind"] == "logprob":
        content, logprobs_content = await _call_judge_llm(session, judge["llm"], prompt, want_logprobs=True)
        if content is None:
            return None
        if not logprobs_content:
            logger.error(
                "Judge '%s': no usable logprobs returned by model %s "
                "(unverified KubeAI/vLLM logprobs support) — skipping cell.",
                judge["name"], judge["llm"],
            )
            return None
        top_logprobs = _locate_score_token(logprobs_content, judge["scale"])
        if top_logprobs is None:
            logger.error(
                "Judge '%s': no scale-matching token found in response (scanned %d tokens, "
                "post-</think> aware) — skipping cell.",
                judge["name"], len(logprobs_content),
            )
            return None
        score = _score_from_logprobs(top_logprobs, judge["scale"])
        if score is None:
            logger.error(
                "Judge '%s': matched token had no scale-matching alternatives in top_logprobs "
                "(%r) — skipping cell.",
                judge["name"], top_logprobs,
            )
        return score

    content, _ = await _call_judge_llm(session, judge["llm"], prompt, want_logprobs=False)
    if content is None:
        return None
    score = _score_from_label(content, judge["scale"])
    if score is None:
        logger.error("Judge '%s': could not parse a valid scale value from reply %r — skipping cell.", judge["name"], content)
    return score


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------


async def main() -> None:
    """Main sync loop: scan for sidecars → wait for transcripts → LLM → upload."""
    load_dotenv()

    client, remote_roots = _make_webdav_client()

    tmp_dir = Path(tempfile.mkdtemp(prefix="nc_llm_"))
    try:
        logger.info("Scanning %s for sidecar-driven LLM jobs...", remote_roots)
        jobs, batch_folders, folder_roots = _collect_llm_jobs(client, remote_roots, tmp_dir)

        # Defined before the branch below (not inside `else:`) — a true
        # no-op run (no new jobs) still needs this defined for the table-
        # rebuild gate right after, and for the judge pass, which can have
        # work even when this run found zero new LLM jobs.
        updated_llm_dirs: set[str] = set()

        if not jobs:
            logger.info("No LLM jobs ready to run.")
        else:
            logger.info("Found %d LLM job(s) to run.", len(jobs))

            # Group by model (stable sort keeps each model's jobs in their
            # original crawl-discovery order) so consecutive calls to the same
            # KubeAI-hosted model land within its scaleDownDelaySeconds warm
            # window instead of round-robin-ing across models and forcing a
            # cold start on every single call.
            jobs.sort(key=lambda j: j["model"])
            logger.info("Processing order (grouped by model): %s", [j["model"] for j in jobs][:50])

            prompt_cache: dict[str, str] = {}
            timeout = aiohttp.ClientTimeout(total=int(os.environ.get("LLM_TIMEOUT", "900")))
            async with aiohttp.ClientSession(timeout=timeout) as session:
                for job in tqdm(jobs, desc="Processing"):
                    try:
                        transcript = _download_text(client, job["transcript_path"], tmp_dir)
                    except Exception as exc:
                        logger.error("Failed to download transcript %s: %s", job["transcript_path"], exc)
                        continue

                    try:
                        template = _load_prompt_template(
                            client, job["remote_root"], job["prompt"], tmp_dir, prompt_cache
                        )
                    except FileNotFoundError as exc:
                        logger.error("%s Skipping %s.", exc, job["transcript_path"])
                        continue
                    prompt = _render_prompt(template, transcript, job["context"], job["extra_fields"])

                    answer = await _call_llm(session, job["model"], prompt)
                    if answer is None:
                        continue

                    try:
                        client.list(job["llm_dir"])
                    except RemoteResourceNotFound:
                        client.mkdir(job["llm_dir"])

                    local_out = tmp_dir / job["output_name"]
                    local_out.write_text(answer, encoding="utf-8")
                    remote_out = job["llm_dir"] + job["output_name"]
                    try:
                        client.upload_sync(local_path=str(local_out), remote_path=remote_out)
                        logger.info("Uploaded %s", remote_out)
                        updated_llm_dirs.add(job["llm_dir"])
                    except Exception as exc:
                        logger.error("Upload failed for %s: %s", remote_out, exc)
                    local_out.unlink(missing_ok=True)

        # Only rebuild a folder's per-prompt review tables if a job actually
        # uploaded something new into its llm/ this run — a table is a
        # snapshot of llm/, and llm/ didn't change otherwise. Without this,
        # every cron tick re-downloaded and re-uploaded every table even on
        # a fully-processed, no-op run.
        for folder_path, rows in batch_folders.items():
            llm_dir = folder_path.rstrip("/") + f"/{LLM_SUBFOLDER}/"
            if llm_dir in updated_llm_dirs:
                _build_prompt_tables(client, folder_path, rows, tmp_dir)

        # Judge pass — independent of the jobs loop above (judges may have
        # work even on a run with zero new LLM jobs, e.g. outputs from a
        # prior run still awaiting judging). File-driven kill switch: unset
        # or "true" (default) runs it, any other value disables it cluster-
        # wide without needing to delete every judges/*.yaml.
        if os.environ.get("LLM_JUDGE_ENABLED", "true").lower() != "false":
            judges_by_root = _collect_judges(client, remote_roots, tmp_dir)
            if judges_by_root:
                groups = _collect_judge_scoring_groups(client, batch_folders, folder_roots, judges_by_root, tmp_dir)
                cell_jobs = [
                    {
                        **group,
                        "prompt": prompt,
                        "stem": stem,
                        "model": model,
                        "output_name": filename,
                        "existing_scores": pdata["existing_scores"],
                    }
                    for group in groups
                    for prompt, pdata in group["prompts"].items()
                    for stem, model, filename in _missing_judge_cells(pdata["cells"], pdata["existing_scores"])
                ]
                if cell_jobs:
                    # Same warm-model-grouping discipline as the main jobs loop.
                    cell_jobs.sort(key=lambda j: j["judge"]["llm"])
                    logger.info("Found %d judge-scoring cell(s) to run.", len(cell_jobs))

                    # Upload each group's scores.xlsx as soon as ITS cells are
                    # all attempted, not once at the very end — a large
                    # backlog can have hundreds of cells across many groups;
                    # waiting for every single one before any output becomes
                    # visible means nothing shows up for a very long time.
                    remaining: dict[str, int] = {}
                    touched_paths: set[str] = set()
                    for cj in cell_jobs:
                        remaining[cj["scores_remote_path"]] = remaining.get(cj["scores_remote_path"], 0) + 1
                    groups_by_path = {g["scores_remote_path"]: g for g in groups}

                    timeout = aiohttp.ClientTimeout(total=int(os.environ.get("LLM_TIMEOUT", "900")))
                    async with aiohttp.ClientSession(timeout=timeout) as session:
                        for cj in tqdm(cell_jobs, desc="Judging"):
                            score = await _score_one_judge_cell(client, session, cj, tmp_dir)
                            if score is not None:
                                cj["existing_scores"].setdefault(cj["stem"], {})[cj["model"]] = score
                                touched_paths.add(cj["scores_remote_path"])

                            remaining[cj["scores_remote_path"]] -= 1
                            if remaining[cj["scores_remote_path"]] == 0 and cj["scores_remote_path"] in touched_paths:
                                _upload_judge_group_scores(client, groups_by_path[cj["scores_remote_path"]], tmp_dir)

    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)

    logger.info("Done.")


if __name__ == "__main__":
    asyncio.run(main())
